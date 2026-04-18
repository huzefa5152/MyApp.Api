"""Create generic Excel templates with merge field placeholders for Challan, Bill, and Tax Invoice."""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "data", "excel-templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

thin = Side(style="thin")
medium = Side(style="medium")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_medium = Border(left=medium, right=medium, top=medium, bottom=medium)
header_fill = PatternFill("solid", fgColor="366092")
light_fill = PatternFill("solid", fgColor="DCE6F1")
white_fill = PatternFill("solid", fgColor="FFFFFF")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def style_range(ws, row, col_start, col_end, font=None, fill=None, alignment=None, border=None):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        if font: cell.font = font
        if fill: cell.fill = fill
        if alignment: cell.alignment = alignment
        if border: cell.border = border


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 1. DELIVERY CHALLAN TEMPLATE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_challan_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Delivery Challan"
    ws.page_setup.paperSize = 9  # A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    set_col_widths(ws, [3, 14, 12, 18, 12, 10, 22])

    title_font = Font(name="Arial", size=22, bold=True, color="1F4E79")
    subtitle_font = Font(name="Arial", size=14, bold=True, color="366092")
    label_font = Font(name="Arial", size=10, bold=True)
    value_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    item_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrapText=True)

    r = 1
    # Company Name
    ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=5)
    ws.cell(r, 2, "{{companyBrandName}}").font = title_font
    ws.cell(r, 2).alignment = Alignment(vertical="center")
    # DC Title
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    ws.cell(r, 6, "Delivery Challan").font = subtitle_font
    ws.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")

    r = 2
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    ws.cell(r, 6, "{{fmtDate deliveryDate}}").font = value_font
    ws.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")

    r = 3
    ws.merge_cells(start_row=r, start_column=2, end_row=r+1, end_column=5)
    ws.cell(r, 2, "{{companyAddress}}").font = value_font
    ws.cell(r, 2).alignment = left_wrap
    ws.row_dimensions[r].height = 18

    r = 4
    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    ws.cell(r, 6, "DC # {{challanNumber}}").font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    ws.cell(r, 6).alignment = Alignment(horizontal="right", vertical="center")

    r = 5
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(r, 2, "{{companyPhone}}").font = value_font
    ws.cell(r, 2).alignment = left_wrap

    r = 7
    ws.cell(r, 2, "Order Date:").font = label_font
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(r, 3, "{{fmtDate deliveryDate}}").font = value_font
    ws.cell(r, 6, "PO #:").font = label_font
    ws.cell(r, 7, "{{poNumber}}").font = value_font

    r = 8
    ws.cell(r, 2, "PO Date:").font = label_font
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4)
    ws.cell(r, 3, "{{fmtDate poDate}}").font = value_font

    r = 10
    ws.cell(r, 2, "Messers:").font = label_font
    ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=5)
    ws.cell(r, 3, "{{clientName}}").font = Font(name="Arial", size=11, bold=True)
    ws.cell(r, 3).alignment = left_wrap

    r = 12
    ws.cell(r, 2, "Address:").font = label_font
    ws.merge_cells(start_row=r, start_column=3, end_row=r+1, end_column=5)
    ws.cell(r, 3, "{{clientAddress}}").font = value_font
    ws.cell(r, 3).alignment = left_wrap

    r = 14
    ws.cell(r, 2, "Site:").font = label_font
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=5)
    ws.cell(r, 3, "{{clientSite}}").font = value_font

    # Items Table Header
    r = 16
    for c in range(2, 8):
        ws.cell(r, c).fill = header_fill
        ws.cell(r, c).font = header_font
        ws.cell(r, c).alignment = center
        ws.cell(r, c).border = border_all
    ws.cell(r, 2, "S #")
    ws.cell(r, 3, "Qty")
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, "Description")
    ws.cell(r, 7, "Unit")

    # Items Loop
    r = 17
    ws.cell(r, 2, "{{#each items 17}}").font = Font(name="Arial", size=8, color="999999")

    r = 18
    ws.cell(r, 2, "{{@index}}").font = item_font
    ws.cell(r, 2).alignment = center
    ws.cell(r, 2).border = border_all
    ws.cell(r, 3, "{{this.quantity}}").font = item_font
    ws.cell(r, 3).alignment = center
    ws.cell(r, 3).border = border_all
    ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=6)
    ws.cell(r, 4, "{{this.description}}").font = item_font
    ws.cell(r, 4).alignment = left_wrap
    ws.cell(r, 4).border = border_all
    ws.cell(r, 7, "{{this.unit}}").font = item_font
    ws.cell(r, 7).alignment = center
    ws.cell(r, 7).border = border_all
    # borders for merged area
    ws.cell(r, 5).border = border_all
    ws.cell(r, 6).border = border_all

    r = 19
    ws.cell(r, 2, "{{/each}}").font = Font(name="Arial", size=8, color="999999")

    # Footer
    r = 21
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
    ws.cell(r, 2, "Thank you for your business!").font = Font(name="Arial", size=10, italic=True, color="366092")
    ws.cell(r, 2).alignment = center

    r = 24
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    ws.cell(r, 2, "Signature and Stamp").font = Font(name="Arial", size=9, color="666666")
    ws.cell(r, 2).alignment = center
    ws.cell(r, 2).border = Border(top=thin)
    ws.cell(r, 3).border = Border(top=thin)

    ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=7)
    ws.cell(r, 6, "Receiver Signature and Stamp").font = Font(name="Arial", size=9, color="666666")
    ws.cell(r, 6).alignment = center
    ws.cell(r, 6).border = Border(top=thin)
    ws.cell(r, 7).border = Border(top=thin)

    ws.print_area = "A1:G24"
    path = os.path.join(OUTPUT_DIR, "Challan_Template.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 2. BILL / INVOICE TEMPLATE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_bill_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice"
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    set_col_widths(ws, [8, 12, 44, 14, 16])

    title_font = Font(name="Arial", size=20, bold=True, color="1F4E79")
    subtitle_font = Font(name="Arial", size=14, bold=True, color="366092")
    label_font = Font(name="Arial", size=10, bold=True)
    value_font = Font(name="Arial", size=10)
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    item_font = Font(name="Arial", size=10)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrapText=True)

    r = 1
    ws.merge_cells("A1:C1")
    ws.cell(r, 1, "{{companyBrandName}}").font = title_font
    ws.merge_cells("D1:E1")
    ws.cell(r, 4, "BILL").font = subtitle_font
    ws.cell(r, 4).alignment = right

    r = 2
    ws.merge_cells("B2:C3")
    ws.cell(r, 2, "{{companyAddress}}").font = value_font
    ws.cell(r, 2).alignment = left_wrap
    ws.cell(r, 4, "Date:").font = label_font
    ws.cell(r, 5, "{{fmtDate date}}").font = value_font

    r = 3
    ws.merge_cells("D3:E3")
    ws.cell(r, 4, "BILL # {{invoiceNumber}}").font = Font(name="Arial", size=12, bold=True, color="1F4E79")
    ws.cell(r, 4).alignment = right

    r = 4
    ws.merge_cells("A4:B5")
    ws.cell(r, 1, "{{companyPhone}}").font = value_font
    ws.cell(r, 1).alignment = left_wrap

    r = 5
    ws.merge_cells("D5:E5")
    ws.cell(r, 4, "DC # {{join challanNumbers}}").font = value_font
    ws.cell(r, 4).alignment = right

    r = 6
    ws.cell(r, 4, "D.C Date:").font = label_font
    ws.cell(r, 5, "{{joinDates challanDates}}").font = value_font

    r = 8
    ws.cell(r, 1, "To:").font = label_font
    ws.merge_cells("B8:C9")
    ws.cell(r, 2, "{{clientName}}").font = Font(name="Arial", size=11, bold=True)
    ws.cell(r, 2).alignment = left_wrap
    ws.merge_cells("D8:E8")
    ws.cell(r, 4, "NTN # {{clientNTN}}").font = value_font
    ws.cell(r, 4).alignment = right

    r = 9
    ws.merge_cells("D9:E9")
    ws.cell(r, 4, "GST # {{clientSTRN}}").font = value_font
    ws.cell(r, 4).alignment = right

    r = 10
    ws.cell(r, 1, "Address:").font = label_font
    ws.merge_cells("B10:C10")
    ws.cell(r, 2, "{{clientAddress}}").font = value_font
    ws.cell(r, 2).alignment = left_wrap

    r = 11
    ws.cell(r, 1, "Purchase Order #").font = label_font
    ws.cell(r, 3, "{{poNumber}}").font = value_font
    ws.cell(r, 4, "P.O Date:").font = label_font
    ws.cell(r, 5, "{{fmtDate poDate}}").font = value_font

    r = 12
    ws.cell(r, 1, "NTN #").font = label_font
    ws.cell(r, 2, "{{companyNTN}}").font = value_font
    ws.cell(r, 4, "STRN #").font = label_font
    ws.cell(r, 5, "{{companySTRN}}").font = value_font

    # Items Table Header
    r = 14
    headers = ["S #", "Quantity", "Item Details", "Unit Price", "Total Price"]
    for i, h in enumerate(headers, 1):
        c = ws.cell(r, i, h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border_all
    ws.row_dimensions[r].height = 22

    # Items Loop
    r = 15
    ws.cell(r, 1, "{{#each items 20}}").font = Font(name="Arial", size=8, color="999999")

    r = 16
    cols = [("{{this.sNo}}", center), ("{{this.quantity}}", center), ("{{this.description}}", left_wrap), ("{{this.unitPrice}}", right), ("{{this.lineTotal}}", right)]
    for i, (val, align) in enumerate(cols, 1):
        c = ws.cell(r, i, val)
        c.font = item_font
        c.alignment = align
        c.border = border_all

    r = 17
    ws.cell(r, 1, "{{/each}}").font = Font(name="Arial", size=8, color="999999")

    # Totals
    r = 19
    ws.merge_cells("A19:C20")
    ws.cell(r, 1, "Amount In Words:").font = label_font

    ws.cell(r, 4, "SUB TOTAL").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 4).alignment = right
    ws.cell(r, 5, "{{subtotal}}").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 5).alignment = right
    ws.cell(r, 5).border = border_all
    ws.cell(r, 5).number_format = '#,##0'

    r = 20
    ws.cell(r, 4, "GST ({{gstRate}}%)").font = label_font
    ws.cell(r, 4).alignment = right
    ws.cell(r, 5, "{{gstAmount}}").font = value_font
    ws.cell(r, 5).alignment = right
    ws.cell(r, 5).border = border_all
    ws.cell(r, 5).number_format = '#,##0'

    r = 21
    ws.merge_cells("A21:C21")
    ws.cell(r, 1, "{{amountInWords}}").font = Font(name="Arial", size=10, italic=True)
    ws.cell(r, 1).alignment = left_wrap
    ws.cell(r, 4, "GRAND TOTAL").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws.cell(r, 4).alignment = right
    ws.cell(r, 5, "{{grandTotal}}").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws.cell(r, 5).alignment = right
    ws.cell(r, 5).border = border_medium
    ws.cell(r, 5).number_format = '#,##0'

    # Footer
    r = 24
    ws.merge_cells("A24:B24")
    ws.cell(r, 1, "Signature and Stamp").font = Font(name="Arial", size=9, color="666666")
    ws.cell(r, 1).alignment = center
    ws.cell(r, 1).border = Border(top=thin)
    ws.cell(r, 2).border = Border(top=thin)

    ws.merge_cells("D24:E24")
    ws.cell(r, 4, "Receiver Signature and Stamp").font = Font(name="Arial", size=9, color="666666")
    ws.cell(r, 4).alignment = center
    ws.cell(r, 4).border = Border(top=thin)
    ws.cell(r, 5).border = Border(top=thin)

    ws.print_area = "A1:E24"
    path = os.path.join(OUTPUT_DIR, "Bill_Template.xlsx")
    wb.save(path)
    print(f"Created: {path}")


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# 3. SALES TAX INVOICE TEMPLATE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def create_tax_invoice_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Tax Invoice"
    ws.page_setup.paperSize = 9
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3

    set_col_widths(ws, [7, 7, 12, 14, 10, 8, 3, 17, 8, 14, 18])

    title_font = Font(name="Arial", size=18, bold=True, color="1F4E79")
    section_font = Font(name="Arial", size=10, bold=True, color="366092")
    label_font = Font(name="Arial", size=9, bold=True)
    value_font = Font(name="Arial", size=9)
    header_font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    item_font = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center", wrapText=True)
    right = Alignment(horizontal="right", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrapText=True)

    # Title
    r = 2
    ws.merge_cells("A2:K3")
    ws.cell(r, 1, "SALES TAX INVOICE").font = title_font
    ws.cell(r, 1).alignment = center

    r = 5
    ws.merge_cells("A5:B5")
    ws.cell(r, 1, "Invoice No:").font = label_font
    ws.cell(r, 3, "{{invoiceNumber}}").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 4, "Date:").font = label_font
    ws.merge_cells("E5:F5")
    ws.cell(r, 5, "{{fmtDate date}}").font = value_font
    ws.merge_cells("H5:I5")
    ws.cell(r, 8, "DC #:").font = label_font
    ws.merge_cells("J5:K5")
    ws.cell(r, 10, "{{join challanNumbers}}").font = value_font

    # Supplier Section (Left)
    r = 7
    ws.merge_cells("A7:B7")
    ws.cell(r, 1, "Supplier's").font = section_font
    ws.merge_cells("A8:B8")
    ws.cell(r+1, 1, "Name:").font = label_font
    ws.merge_cells("C7:E8")
    ws.cell(r, 3, "{{supplierName}}").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 3).alignment = left_wrap

    # Buyer Section (Right)
    ws.cell(r, 8, "Buyer's").font = section_font
    ws.cell(r+1, 8, "Name:").font = label_font
    ws.merge_cells("I7:K8")
    ws.cell(r, 9, "{{buyerName}}").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 9).alignment = left_wrap

    r = 9
    ws.merge_cells("A9:B9")
    ws.cell(r, 1, "Address:").font = label_font
    ws.merge_cells("C9:E10")
    ws.cell(r, 3, "{{supplierAddress}}").font = value_font
    ws.cell(r, 3).alignment = left_wrap
    ws.cell(r, 8, "Address:").font = label_font
    ws.merge_cells("I9:K10")
    ws.cell(r, 9, "{{buyerAddress}}").font = value_font
    ws.cell(r, 9).alignment = left_wrap

    r = 11
    ws.merge_cells("A11:B11")
    ws.cell(r, 1, "Phone:").font = label_font
    ws.merge_cells("C11:E11")
    ws.cell(r, 3, "{{supplierPhone}}").font = value_font
    ws.cell(r, 8, "Phone:").font = label_font
    ws.merge_cells("I11:K11")
    ws.cell(r, 9, "{{buyerPhone}}").font = value_font

    r = 12
    ws.merge_cells("A12:B12")
    ws.cell(r, 1, "STRN #:").font = label_font
    ws.merge_cells("C12:E12")
    ws.cell(r, 3, "{{supplierSTRN}}").font = value_font
    ws.cell(r, 8, "STRN #:").font = label_font
    ws.merge_cells("I12:K12")
    ws.cell(r, 9, "{{buyerSTRN}}").font = value_font

    r = 13
    ws.merge_cells("A13:B13")
    ws.cell(r, 1, "NTN #:").font = label_font
    ws.merge_cells("C13:E13")
    ws.cell(r, 3, "{{supplierNTN}}").font = value_font
    ws.cell(r, 8, "NTN #:").font = label_font
    ws.merge_cells("I13:K13")
    ws.cell(r, 9, "{{buyerNTN}}").font = value_font

    # PO Info
    r = 15
    ws.merge_cells("A15:K15")
    ws.cell(r, 1, "Purchase Order #: {{poNumber}}").font = value_font

    # Items Table Header
    r = 17
    headers = [("A", "Qty"), ("B", "Unit"), ("C:G", "Description"),
               ("H", "Value Excl.\nSales Tax"), ("I", "Rate of\nSales Tax"),
               ("J", "Total Sales\nTax Payable"), ("K", "Value Incl.\nSales Tax")]
    ws.row_dimensions[r].height = 30

    ws.cell(r, 1, "Qty").font = header_font
    ws.cell(r, 1).fill = header_fill
    ws.cell(r, 1).alignment = center
    ws.cell(r, 1).border = border_all

    ws.cell(r, 2, "Unit").font = header_font
    ws.cell(r, 2).fill = header_fill
    ws.cell(r, 2).alignment = center
    ws.cell(r, 2).border = border_all

    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    ws.cell(r, 3, "Description").font = header_font
    ws.cell(r, 3).fill = header_fill
    ws.cell(r, 3).alignment = center
    ws.cell(r, 3).border = border_all
    for c in range(4, 8):
        ws.cell(r, c).fill = header_fill
        ws.cell(r, c).border = border_all

    ws.cell(r, 8, "Value Excl.\nSales Tax").font = header_font
    ws.cell(r, 8).fill = header_fill
    ws.cell(r, 8).alignment = center
    ws.cell(r, 8).border = border_all

    ws.cell(r, 9, "Rate of\nSales Tax").font = header_font
    ws.cell(r, 9).fill = header_fill
    ws.cell(r, 9).alignment = center
    ws.cell(r, 9).border = border_all

    ws.cell(r, 10, "Total Sales\nTax Payable").font = header_font
    ws.cell(r, 10).fill = header_fill
    ws.cell(r, 10).alignment = center
    ws.cell(r, 10).border = border_all

    ws.cell(r, 11, "Value Incl.\nSales Tax").font = header_font
    ws.cell(r, 11).fill = header_fill
    ws.cell(r, 11).alignment = center
    ws.cell(r, 11).border = border_all

    # Items Loop
    r = 18
    ws.cell(r, 1, "{{#each items 23}}").font = Font(name="Arial", size=8, color="999999")

    r = 19
    item_cols = [
        (1, "{{this.quantity}}", center),
        (2, "{{this.uom}}", center),
        (3, "{{this.description}}", left_wrap),  # merged to col 7
        (8, "{{this.valueExclTax}}", right),
        (9, "{{this.gstRate}}", center),
        (10, "{{this.gstAmount}}", right),
        (11, "{{this.totalInclTax}}", right),
    ]
    for col, val, align in item_cols:
        c = ws.cell(r, col, val)
        c.font = item_font
        c.alignment = align
        c.border = border_all
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)
    for c in range(4, 8):
        ws.cell(r, c).border = border_all

    r = 20
    ws.cell(r, 1, "{{/each}}").font = Font(name="Arial", size=8, color="999999")

    # Totals Row
    r = 22
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(r, 1, "TOTAL:").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 1).alignment = Alignment(horizontal="right", vertical="center")
    ws.cell(r, 1).border = border_all
    for c in range(2, 8):
        ws.cell(r, c).border = border_all

    ws.cell(r, 8, "{{subtotal}}").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 8).alignment = right
    ws.cell(r, 8).border = border_all

    ws.cell(r, 9, "{{gstRate}}%").font = value_font
    ws.cell(r, 9).alignment = center
    ws.cell(r, 9).border = border_all

    ws.cell(r, 10, "{{gstAmount}}").font = Font(name="Arial", size=10, bold=True)
    ws.cell(r, 10).alignment = right
    ws.cell(r, 10).border = border_all

    ws.cell(r, 11, "{{grandTotal}}").font = Font(name="Arial", size=11, bold=True, color="1F4E79")
    ws.cell(r, 11).alignment = right
    ws.cell(r, 11).border = border_medium

    # Amount in words
    r = 24
    ws.merge_cells("A24:D24")
    ws.cell(r, 1, "Amount In Words:").font = label_font
    ws.merge_cells("E24:K25")
    ws.cell(r, 5, "{{amountInWords}}").font = Font(name="Arial", size=10, italic=True)
    ws.cell(r, 5).alignment = left_wrap

    # Footer
    r = 28
    ws.merge_cells("A28:D28")
    ws.cell(r, 1, "SIGNATURE AND STAMP").font = Font(name="Arial", size=9, color="666666")
    ws.cell(r, 1).alignment = center
    ws.cell(r, 1).border = Border(top=thin)
    for c in range(2, 5):
        ws.cell(r, c).border = Border(top=thin)

    ws.merge_cells("I28:K28")
    ws.cell(r, 9, "RECEIVER SIGNATURE AND STAMP").font = Font(name="Arial", size=9, color="666666")
    ws.cell(r, 9).alignment = center
    ws.cell(r, 9).border = Border(top=thin)
    ws.cell(r, 10).border = Border(top=thin)
    ws.cell(r, 11).border = Border(top=thin)

    ws.print_area = "A1:K28"
    path = os.path.join(OUTPUT_DIR, "TaxInvoice_Template.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    create_challan_template()
    create_bill_template()
    create_tax_invoice_template()
    print("\nAll 3 templates created in:", OUTPUT_DIR)
    print("\nMerge fields used:")
    print("  Challan: companyBrandName, companyAddress, companyPhone, challanNumber, deliveryDate, poNumber, poDate, clientName, clientAddress, clientSite, #each items (sNo/@index, quantity, description, unit)")
    print("  Bill: companyBrandName, companyAddress, companyPhone, companyNTN, companySTRN, invoiceNumber, date, challanNumbers, challanDates, poNumber, poDate, clientName, clientAddress, clientNTN, clientSTRN, subtotal, gstRate, gstAmount, grandTotal, amountInWords, #each items (sNo, quantity, description, unitPrice, lineTotal)")
    print("  TaxInvoice: supplierName, supplierAddress, supplierPhone, supplierNTN, supplierSTRN, buyerName, buyerAddress, buyerPhone, buyerNTN, buyerSTRN, invoiceNumber, date, challanNumbers, poNumber, subtotal, gstRate, gstAmount, grandTotal, amountInWords, #each items (quantity, uom, description, valueExclTax, gstRate, gstAmount, totalInclTax)")
