"""
Spreadsheet import suite — layouts, file checks, re-import protection, and the
opening stock importer.

The workbooks are BUILT here rather than read from disk, so the suite runs on
any machine and pins the exact shapes it claims to handle. They reproduce the
real client sheet: a title band, headings on row 3, customs-lot rows from row 4,
HS codes carrying a ":-" suffix, and one item spread across two lots.

What it proves:

  * a layout is saved once, versioned on every rule change, and can be rolled
    back; a rename does not burn a version
  * a private layout is invisible to another company; a shared one is not
  * uploads are checked before parsing — extension, magic bytes, container vs
    extension, opens, non-empty — each with its own message
  * the same workbook fingerprints identically when only the month changes, so
    next month's file matches the same layout
  * preview writes nothing; grouping sums an item held across two lots; the
    match ladder reuses before it creates
  * commit writes item types, opening quantities and the Inventory opening
    balance, and switches the company to tracking inventory
  * the same file cannot be imported twice, a re-saved copy with identical
    content is refused too, a grown sheet still imports, and setting the run
    aside unlocks a deliberate re-import
  * the ledger reconciles every customer against the index sheet before it will
    import, a negative opening becomes a receipt rather than a negative invoice,
    one reference across several rows becomes one invoice, and an undated row
    inherits the date above it
  * imported balances read back through the app's own Customer Ledger, and the
    receivable total lands on Accounts receivable with the ledger frozen
  * a built-in layout ships for each kind, cannot be deleted, and is offered as
    a starting point when a workbook is not recognised
  * a layout keeps recognising its own template after every value in the file
    has changed — the property that stops a monthly re-upload being re-mapped
  * a period is supplied per import, so a layout carries no dates and stays
    correct year on year
  * every route refuses a company the caller cannot reach

    python scripts/test_spreadsheet_import.py --base http://localhost:5134

Creates its own throwaway companies and deletes them at the end unless --keep.
"""

import argparse
import io
import json
import sys
import uuid
from datetime import date, datetime

import requests

try:
    import openpyxl
except ImportError:
    print("openpyxl is required:  pip install openpyxl")
    sys.exit(2)

# Messages echoed from the API carry typographic characters; a Windows console
# defaults to cp1252 and would raise on them mid-suite.
for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

PASS, FAIL, SKIP = "PASS", "FAIL", "SKIP"
results = []


def check(name, ok, detail=""):
    results.append((PASS if ok else FAIL, name, detail))
    print(f"[{PASS if ok else FAIL}] {name}" + (f" — {detail}" if detail else ""))
    return ok


def skip(name, why):
    results.append((SKIP, name, why))
    print(f"[{SKIP}] {name} — {why}")


def login(base, username, password):
    r = requests.post(f"{base}/api/auth/login",
                      json={"username": username, "password": password}, timeout=30)
    r.raise_for_status()
    return r.json()["token"]


# ── Workbook builders ───────────────────────────────────────────────────────
# Mirrors the client's real stock sheet: a title band, headings on row 3, data
# from row 4, HS codes with a ":-" tail, and the Balance block at cols 18/19.

STOCK_MAPPING = {
    "sheetSelect": {"mode": "byHeaderText", "mustContain": ["GD Number"]},
    "headerRow": 3,
    "firstDataRow": 4,
    "columns": {
        "lotRef": 2, "lotDate": 3, "hsCodeShort": 4, "hsCodeFull": 5,
        "itemName": 6, "unit": 9, "balanceQty": 18, "balanceValue": 19,
    },
    "hsCodeStripSuffix": ":-",
    "ignoreColumns": [7],
}


def stock_workbook(rows, month="Jul 2026", sheet_name=None):
    """rows: (lot, hs4, hs8, name, subcat, unit, qty, value)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name or month
    ws.cell(1, 6, "ALPHA Trader")
    ws.cell(2, 1, "Stock Sheet")
    ws.cell(2, 6, month)
    headings = {1: "Claimed Month", 2: "GD Number", 3: "GD Date", 4: "4 Digit Hs Code",
                5: "8 Digit Hs code", 6: "Items", 7: "Sub Catory", 8: "Price",
                9: "Unit", 10: "Qty", 11: "Excluding", 18: " Qty", 19: "Balance Exl"}
    for col, text in headings.items():
        ws.cell(3, col, text)

    for i, (lot, hs4, hs8, name, subcat, unit, qty, value) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, month)
        ws.cell(r, 2, lot)
        ws.cell(r, 3, "23-07-2025")
        ws.cell(r, 4, hs4)
        ws.cell(r, 5, hs8)
        ws.cell(r, 6, name)
        ws.cell(r, 7, subcat)
        ws.cell(r, 9, unit)
        ws.cell(r, 18, qty)
        ws.cell(r, 19, value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# Every HS code below is a REAL Pakistan tariff line. Validation is master-first,
# so an invented code is rejected — and Pakistan splits some WCO subheadings into
# national lines, which is why this uses 8536.5010 rather than 8536.5000.
def base_rows(tag):
    """Two lots of one item, plus three singles — 5 rows, 4 items."""
    return [
        ("KAPE-HC-5876", "8481", "8481.1000:-", f"SS BALL VALVE {tag}", "Valve", "Pcs", 12, 23160),
        ("KAPE-HC-5876", "8513", "8513.1090:-", f"RECHARGEABLE LED {tag}", "Electrical", "Pcs", 46, 30000),
        ("KAPE-HC-5876", "8450", "8450.9000:-", f"WASHING PARTS {tag}", "Electrical", "Kg", 67.5, 54160),
        # Same item, second customs lot — must merge into one item type.
        ("KAPW-HC-128753", "8450", "8450.9000:-", f"WASHING PARTS {tag}", "Electrical", "KG", 32.5, 26000),
        ("KAPW-HC-128753", "8712", "8712.0000:-", f"CHILDREN BICYCLE {tag}", "Sports", "Pcs", 2, 15120),
    ]



LEDGER_MAPPING = {
    "indexSheet": {"mode": "byName", "name": "Index"},
    "indexFirstRow": 3,
    "indexColumns": {"name": 2, "opening": 3, "debit": 4, "credit": 5, "closing": 6},
    "clientSheets": {"mode": "allExcept", "except": ["Index"]},
    "clientNameCell": "A3",
    "firstDataRow": 6,
    "columns": {"date": 2, "refAny": [3, 4], "debit": 6, "credit": 7, "balance": 8},
    "creditIsInvoice": True,
    "refPattern": r"^[A-Za-z]{1,4}-\d+$",
    "undatedRule": "carryPreviousRow",
    "openingDate": "2025-06-30",
    "periodStart": "2025-07-01",
    "periodEnd": "2026-06-30",
    "openingBand": 900000,
    "unreferencedBand": 950000,
}


def ledger_workbook(clients):
    """clients: list of dicts with name, tab, opening, rows.

    rows: (date|None, ref_col_c, particulars_col_d, debit, credit)
    Mirrors the real workbook: an index sheet, then one sheet per customer with
    the customer's name in A3 and transactions from row 6.
    """
    wb = openpyxl.Workbook()
    idx = wb.active
    idx.title = "Index"
    idx.cell(1, 1, "ALPHA TRADERS")
    for col, text in {1: "S.No", 2: "Name", 3: "Opening Balance",
                      4: "Debit", 5: "Credit", 6: "Closing Balance"}.items():
        idx.cell(2, col, text)

    for i, c in enumerate(clients):
        debit = sum(r[3] for r in c["rows"])
        credit = sum(r[4] for r in c["rows"])
        row = 3 + i
        idx.cell(row, 1, i + 1)
        idx.cell(row, 2, c["name"])
        idx.cell(row, 3, c["opening"])
        idx.cell(row, 4, debit)
        idx.cell(row, 5, credit)
        idx.cell(row, 6, c["opening"] + credit - debit)

    for c in clients:
        ws = wb.create_sheet(c["tab"])
        ws.cell(1, 1, "ALPHA TRADERS")
        ws.cell(2, 1, "Ledger")
        ws.cell(3, 1, c.get("sheetName", c["name"]))
        for col, text in {1: "S.No", 2: "Date", 3: "Inv", 4: "Particulars",
                          5: "Opening", 6: "Debit", 7: "Credit", 8: "Balance"}.items():
            ws.cell(5, col, text)
        running = c["opening"]
        for i, (dt, ref_c, ref_d, debit, credit) in enumerate(c["rows"]):
            r = 6 + i
            ws.cell(r, 1, i + 1)
            if dt:
                ws.cell(r, 2, dt)
            if ref_c:
                ws.cell(r, 3, ref_c)
            if ref_d:
                ws.cell(r, 4, ref_d)
            if debit:
                ws.cell(r, 6, debit)
            if credit:
                ws.cell(r, 7, credit)
            running += credit - debit
            ws.cell(r, 8, running)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def empty_workbook():
    wb = openpyxl.Workbook()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload(url, h, content, filename="stock.xlsx", data=None, params=None):
    return requests.post(
        url, params=params or {}, data=data or {},
        files={"file": (filename, io.BytesIO(content),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=h, timeout=180)


def make_company(api, h, name):
    r = requests.post(f"{api}/companies", headers=h, timeout=60, json={
        "name": name, "brandName": "SSIMP", "fullAddress": "1 Test Street",
        "phone": "021-0000000", "ntn": "1234567-8",
        "startingChallanNumber": 1, "startingInvoiceNumber": 1,
        "startingSalesQuoteNumber": 1, "startingSalesOrderNumber": 1,
    })
    if r.status_code not in (200, 201):
        raise RuntimeError(f"company create failed: http {r.status_code} {r.text[:200]}")
    return r.json()["id"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default="http://localhost:5134")
    ap.add_argument("--username", default="admin")
    ap.add_argument("--password", default="admin123")
    ap.add_argument("--keep", action="store_true")
    args = ap.parse_args()

    base = args.base.rstrip("/")
    api = f"{base}/api"
    h = {"Authorization": f"Bearer {login(base, args.username, args.password)}"}

    tag = uuid.uuid4().hex[:6].upper()
    company = make_company(api, h, f"Sheet Import Co {tag}")
    other = make_company(api, h, f"Sheet Import Other {tag}")
    profile_ids = []

    stock_preview = f"{api}/spreadsheet-import/opening-stock/preview"
    stock_commit = f"{api}/spreadsheet-import/opening-stock/commit"

    try:
        requests.post(f"{api}/accounts/company/{company}/seed-wholesale", headers=h, timeout=120)

        # ── 1. Layouts ──────────────────────────────────────────────────
        r = requests.post(f"{api}/import-profiles", headers=h, timeout=30, json={
            "kind": "OpeningStock", "layout": "LotRows", "name": f"Stock {tag}",
            "companyId": company, "signatureHash": "a" * 64,
            "tokenSignature": "balance|excluding|items|qty",
            "mappingJson": json.dumps(STOCK_MAPPING),
        })
        ok = r.status_code in (200, 201)
        profile = r.json() if ok else {}
        if ok:
            profile_ids.append(profile["id"])
        check("a layout saves as version 1", ok and profile.get("currentVersion") == 1,
              f"http {r.status_code}")
        check("a company-scoped layout is not shared", profile.get("isShared") is False)

        pid = profile.get("id")

        r = requests.put(f"{api}/import-profiles/{pid}", headers=h, timeout=30,
                         json={"name": f"Stock {tag} renamed"})
        check("a rename does not burn a version",
              r.ok and r.json().get("currentVersion") == 1,
              f"version={r.json().get('currentVersion') if r.ok else '-'}")

        changed = dict(STOCK_MAPPING, firstDataRow=5)
        r = requests.put(f"{api}/import-profiles/{pid}", headers=h, timeout=30,
                         json={"mappingJson": json.dumps(changed), "changeNote": "start row"})
        check("a mapping change bumps the version",
              r.ok and r.json().get("currentVersion") == 2,
              f"version={r.json().get('currentVersion') if r.ok else '-'}")

        r = requests.get(f"{api}/import-profiles/{pid}/versions", headers=h, timeout=30)
        check("both versions are in the history", r.ok and len(r.json()) == 2,
              f"{len(r.json()) if r.ok else '-'} versions")

        r = requests.post(f"{api}/import-profiles/{pid}/rollback", headers=h, timeout=30,
                          json={"version": 1})
        rolled = r.json() if r.ok else {}
        check("rollback restores the old mapping",
              r.ok and json.loads(rolled.get("mappingJson", "{}")).get("firstDataRow") == 4,
              f"http {r.status_code}")
        check("rollback moves history forward, never back",
              rolled.get("currentVersion") == 3, f"version={rolled.get('currentVersion')}")

        r = requests.post(f"{api}/import-profiles", headers=h, timeout=30, json={
            "kind": "OpeningStock", "layout": "NotALayout", "name": "bad",
            "companyId": company, "signatureHash": "b" * 64, "mappingJson": "{}",
        })
        check("an unknown layout is refused", r.status_code == 400, f"http {r.status_code}")

        r = requests.post(f"{api}/import-profiles", headers=h, timeout=30, json={
            "kind": "OpeningStock", "layout": "LotRows", "name": "bad json",
            "companyId": company, "signatureHash": "c" * 64, "mappingJson": "{not json",
        })
        check("a malformed mapping is refused", r.status_code == 400, f"http {r.status_code}")

        r = requests.get(f"{api}/import-profiles", headers=h, timeout=30,
                         params={"kind": "OpeningStock", "companyId": other})
        visible = [p["id"] for p in r.json()] if r.ok else []
        check("a private layout is invisible to another company", pid not in visible,
              f"{len(visible)} visible to the other company")

        # ── 2. File validation ──────────────────────────────────────────
        good = stock_workbook(base_rows(tag))
        params = {"companyId": company}
        form = {"mappingJson": json.dumps(STOCK_MAPPING)}

        r = upload(stock_preview, h, b"", "stock.xlsx", form, params)
        check("an empty file is refused", r.status_code == 400, f"http {r.status_code}")

        r = upload(stock_preview, h, good, "stock.txt", form, params)
        check("a .txt is refused on extension", r.status_code == 400, f"http {r.status_code}")

        r = upload(stock_preview, h, b"%PDF-1.4\n%fake pdf bytes here", "stock.xlsx", form, params)
        msg = (r.json().get("message") or "") if r.status_code == 400 else ""
        check("a PDF renamed .xlsx is refused on magic bytes",
              r.status_code == 400 and "not a valid Excel" in msg, f"http {r.status_code}: {msg[:60]}")

        # OLE2 header = a real legacy .xls container, wrong for a .xlsx name.
        r = upload(stock_preview, h, bytes([0xD0, 0xCF, 0x11, 0xE0, 0xA1, 0xB1, 0x1A, 0xE1]) + b"\x00" * 64,
                   "stock.xlsx", form, params)
        msg = (r.json().get("message") or "") if r.status_code == 400 else ""
        check("a .xls renamed .xlsx is refused on the container",
              r.status_code == 400 and "do not match" in msg, f"http {r.status_code}: {msg[:60]}")

        r = upload(stock_preview, h, empty_workbook(), "stock.xlsx", form, params)
        msg = (r.json().get("message") or "") if r.status_code == 400 else ""
        check("an empty workbook is refused", r.status_code == 400 and "empty" in msg.lower(),
              f"http {r.status_code}: {msg[:60]}")

        # ── 3. Fingerprint ──────────────────────────────────────────────
        idurl = f"{api}/spreadsheet-import/identify"
        r = upload(idurl, h, good, "stock.xlsx", None, {"companyId": company, "kind": "OpeningStock"})
        ident = r.json() if r.ok else {}
        sig_jul = ident.get("signatureHash")
        check("identify fingerprints the workbook", r.ok and bool(sig_jul), f"http {r.status_code}")
        check("identify describes the sheets for mapping", len(ident.get("sheets", [])) >= 1,
              f"{len(ident.get('sheets', []))} sheets")

        aug = stock_workbook(base_rows(tag), month="Aug 2026")
        r = upload(idurl, h, aug, "stock.xlsx", None, {"companyId": company, "kind": "OpeningStock"})
        check("next month's file fingerprints the same",
              r.ok and r.json().get("signatureHash") == sig_jul,
              "digits are stripped, so the month does not change the layout")

        # ── 4. Opening stock preview ────────────────────────────────────
        r = upload(stock_preview, h, good, "stock.xlsx", form, params)
        prev = r.json() if r.ok else {}
        check("the stock sheet previews", r.ok, f"http {r.status_code}: {r.text[:160]}")
        check("5 sheet rows become 4 items", prev.get("sourceRowCount") == 5 and len(prev.get("rows", [])) == 4,
              f"rows={prev.get('sourceRowCount')} items={len(prev.get('rows', []))}")

        merged = next((x for x in prev.get("rows", []) if "WASHING PARTS" in x["itemName"]), None)
        check("an item held across two lots sums its quantities",
              merged is not None and abs(merged["quantity"] - 100.0) < 0.001,
              f"qty={merged['quantity'] if merged else '-'}")
        check("both lot references are kept on the merged item",
              merged is not None and merged.get("lotRefs", "").count(",") == 1,
              f"lots={merged.get('lotRefs') if merged else '-'}")
        check("the sub-category column is ignored",
              all("Electrical" not in (x.get("itemName") or "") for x in prev.get("rows", [])))

        # Two DIFFERENT product names under ONE tariff code. An item type is
        # identified by its code, so both rows resolve to the same catalog row
        # at commit time -- they have to be ADDED. Grouping by name instead
        # gave one preview row each and the last one written overwrote its
        # sibling's opening balance, which silently lost 11 items and
        # 2,312,996.50 of value from the client's real sheet (2026-09-02).
        shared_rows = [
            ("SHARED-1", "8513", "8513.1090:-", f"LED DISPLAY {tag}", "Electrical", "Pcs", 46, 30000),
            ("SHARED-2", "8513", "8513.1090:-", f"LED CANDLE {tag}", "Electrical", "Pcs", 110, 71191),
            ("SHARED-3", "8513", "8513.1090:-", f"LED CEILING {tag}", "Electrical", "Pcs", 150, 107274),
        ]
        r = upload(stock_preview, h, stock_workbook(shared_rows), "stock.xlsx", form, params)
        ps = r.json() if r.ok else {}
        check("three rows under one HS code become one item",
              len(ps.get("rows", [])) == 1,
              f"got {len(ps.get('rows', []))} items")
        one = (ps.get("rows") or [{}])[0]
        check("their quantities are added together",
              abs(float(one.get("quantity") or 0) - 306) < 0.001,
              f"qty={one.get('quantity')}")
        check("their values are added together",
              abs(float(one.get("value") or 0) - 208465) < 0.5,
              f"value={one.get('value')}")
        check("the total still equals the sheet",
              abs(float(ps.get("totalValue") or 0) - 208465) < 0.5,
              f"total={ps.get('totalValue')}")
        check("the merged row says which names it folded in",
              any("added together with" in m for m in (one.get("messages") or [])),
              f"messages={one.get('messages')}")

        # A row with NO code still groups on its name, so two unrelated
        # un-coded items cannot collapse into each other.
        noco = [
            ("NC-1", "", "", f"NO CODE A {tag}", "X", "Pcs", 5, 500),
            ("NC-2", "", "", f"NO CODE B {tag}", "X", "Pcs", 7, 700),
        ]
        r = upload(stock_preview, h, stock_workbook(noco), "stock.xlsx", form, params)
        pn = r.json() if r.ok else {}
        check("two un-coded items stay separate",
              len(pn.get("rows", [])) == 2, f"got {len(pn.get('rows', []))}")
        check("the HS code suffix is stripped",
              all((x.get("hsCode") or "").replace(".", "").isdigit() for x in prev.get("rows", [])),
              f"codes={[x.get('hsCode') for x in prev.get('rows', [])][:4]}")
        check("every item is new on a first upload",
              prev.get("statusCounts", {}).get("will-create") == 4,
              f"counts={prev.get('statusCounts')}")
        check("the value total adds up",
              abs(prev.get("totalValue", 0) - 148440) < 0.5, f"value={prev.get('totalValue')}")

        before = requests.get(f"{api}/stock/company/{company}/opening", headers=h, timeout=60)
        check("preview wrote nothing", before.ok and len(before.json()) == 0,
              f"{len(before.json()) if before.ok else '-'} opening balances exist")

        hs_count = requests.get(f"{api}/hscodes/count", headers=h, timeout=30)
        master_loaded = hs_count.ok and (hs_count.json().get("count", 0) if isinstance(hs_count.json(), dict) else hs_count.json()) > 0
        if master_loaded:
            bogus = stock_workbook([("L1", "9999", "9999.9999:-", f"BOGUS {tag}", "X", "Pcs", 1, 10)])
            r = upload(stock_preview, h, bogus, "stock.xlsx", form, params)
            p2 = r.json() if r.ok else {}
            check("an HS code missing from the master blocks the row",
                  p2.get("statusCounts", {}).get("hs-unknown") == 1 and not p2.get("canCommit"),
                  f"counts={p2.get('statusCounts')}")
        else:
            skip("an HS code missing from the master blocks the row",
                 "HS master is empty on this database; run the HS Code import first")
            check("an empty HS master is reported, not silently ignored",
                  any("HS code master is empty" in w for w in prev.get("warnings", [])),
                  f"warnings={prev.get('warnings')}")

        # ── 5. Commit ───────────────────────────────────────────────────
        body = {
            "companyId": company,
            "fileSha256": prev["fileSha256"],
            "fileName": "stock.xlsx",
            "fileSizeBytes": prev["fileSizeBytes"],
            "asOfDate": date(2026, 7, 1).isoformat(),
            "postInventoryValue": True,
            "enableInventoryTracking": True,
            "rows": [
                {"itemName": x["itemName"], "hsCode": x["hsCode"],
                 "isHsCodePartial": x["isHsCodePartial"], "unit": x["unit"],
                 "quantity": x["quantity"], "value": x["value"],
                 "lotRefs": x["lotRefs"], "itemTypeId": x["itemTypeId"]}
                for x in prev["rows"]
            ],
        }
        r = requests.post(stock_commit, headers=h, timeout=300, json=body)
        res = r.json() if r.ok else {}
        check("commit succeeds", r.ok, f"http {r.status_code}: {r.text[:200]}")
        check("four item types are created", res.get("itemTypesCreated") == 4,
              f"created={res.get('itemTypesCreated')}")
        check("four opening balances are written", res.get("openingBalancesWritten") == 4,
              f"written={res.get('openingBalancesWritten')}")
        check("the stock value is posted to Inventory",
              abs(res.get("inventoryValuePosted", 0) - 148440) < 0.5,
              f"posted={res.get('inventoryValuePosted')}")
        check("an import run is recorded", res.get("importRunId", 0) > 0)

        after = requests.get(f"{api}/stock/company/{company}/opening", headers=h, timeout=60).json()
        check("the opening balances landed", len(after) == 4, f"{len(after)} rows")
        washing = next((x for x in after if "WASHING PARTS" in x["itemTypeName"]), None)
        check("the merged item carries the summed quantity",
              washing is not None and abs(washing["quantity"] - 100.0) < 0.001,
              f"qty={washing['quantity'] if washing else '-'}")
        check("the lot references are on the opening balance note",
              washing is not None and "KAPE-HC-5876" in (washing.get("notes") or ""),
              f"notes={washing.get('notes') if washing else '-'}")

        r = requests.get(f"{api}/accounts/company/{company}/flat", headers=h, timeout=60)
        # ControlType serialises as its NAME, not the enum's number.
        inv = next((a for a in r.json() if a.get("controlType") == "Inventory"), None) if r.ok else None
        check("the Inventory account carries the opening balance",
              inv is not None and abs(inv.get("openingBalance", 0) - 148440) < 0.5,
              f"opening={inv.get('openingBalance') if inv else '-'}")

        r = requests.get(f"{api}/companies/{company}", headers=h, timeout=30)
        comp = r.json() if r.ok else {}
        check("the company now tracks inventory", comp.get("inventoryTrackingEnabled") is True)
        check("the company is on the standard inventory flow",
              comp.get("inventoryFlowVersion") == 2, f"version={comp.get('inventoryFlowVersion')}")

        # ── 6. Re-import protection ─────────────────────────────────────
        r = upload(stock_preview, h, good, "stock.xlsx", form, params)
        p3 = r.json() if r.ok else {}
        check("the same file is refused on a second preview",
              any("already imported" in e for e in p3.get("blockingErrors", [])),
              f"errors={p3.get('blockingErrors')}")

        r = requests.post(stock_commit, headers=h, timeout=120, json=body)
        check("the same file is refused at commit too", r.status_code == 400,
              f"http {r.status_code}")

        counts = requests.get(f"{api}/stock/company/{company}/opening", headers=h, timeout=60).json()
        check("the refused re-import wrote nothing", len(counts) == 4, f"{len(counts)} rows")

        # A re-save produces different bytes and identical content: the hash
        # misses it, so the content check has to catch it.
        resaved = stock_workbook(base_rows(tag), sheet_name="Jul 2026 ")
        r = upload(stock_preview, h, resaved, "stock-resaved.xlsx", form, params)
        p4 = r.json() if r.ok else {}
        check("a re-saved copy has a different hash",
              p4.get("fileSha256") != prev["fileSha256"], "bytes differ, as expected")
        check("a re-saved copy with identical content is still refused",
              any("already been imported" in e for e in p4.get("blockingErrors", [])),
              f"errors={p4.get('blockingErrors')}")

        # A sheet that has genuinely grown must still import.
        grown = stock_workbook(base_rows(tag) + [
            ("KAPE-HC-9000", "8536", "8536.5010:-", f"NEW SWITCH {tag}", "Electrical", "Pcs", 25, 5000)])
        r = upload(stock_preview, h, grown, "stock-grown.xlsx", form, params)
        p5 = r.json() if r.ok else {}
        check("a sheet with new rows still imports", p5.get("canCommit") is True,
              f"errors={p5.get('blockingErrors')}")
        check("the unchanged rows are called out, not hidden",
              any("already carry exactly this opening quantity" in w for w in p5.get("warnings", [])),
              f"warnings={p5.get('warnings')}")

        # ── 7. Force re-import ──────────────────────────────────────────
        runs = requests.get(f"{api}/spreadsheet-import/runs", headers=h, timeout=30,
                            params={"companyId": company}).json()
        check("the import history lists the run", runs.get("totalCount", 0) >= 1,
              f"total={runs.get('totalCount')}")
        run_id = runs["items"][0]["id"]

        r = requests.post(f"{api}/spreadsheet-import/runs/{run_id}/supersede", headers=h, timeout=30,
                          params={"companyId": company}, json={"reason": ""})
        check("setting a run aside needs a reason", r.status_code == 400, f"http {r.status_code}")

        r = requests.post(f"{api}/spreadsheet-import/runs/{run_id}/supersede", headers=h, timeout=30,
                          params={"companyId": company}, json={"reason": "wrong file"})
        check("a run can be set aside with a reason", r.ok and r.json().get("isSuperseded") is True,
              f"http {r.status_code}")

        r = upload(stock_preview, h, good, "stock.xlsx", form, params)
        p6 = r.json() if r.ok else {}
        check("the same file may be previewed again once set aside",
              not any("already imported on" in e for e in p6.get("blockingErrors", [])),
              f"errors={p6.get('blockingErrors')}")

        # ── 9. Customer ledger ──────────────────────────────────────────
        ledger_preview = f"{api}/spreadsheet-import/customer-ledger/preview"
        ledger_commit = f"{api}/spreadsheet-import/customer-ledger/commit"
        lparams = {"companyId": other}
        lform = {"mappingJson": json.dumps(LEDGER_MAPPING)}

        d1, d2, d3 = datetime(2025, 8, 4), datetime(2025, 9, 12), datetime(2026, 2, 3)
        clients_spec = [
            # Ordinary customer: opening, three invoices, one receipt.
            {"name": f"Alpha Hardware {tag}", "tab": f"ALPHA HW {tag}", "opening": 100000.0,
             "rows": [(d1, None, "AA-1", 0, 50000.0),
                      (d2, None, "AA-2", 0, 25000.0),
                      (None, None, "Cash Rec", 30000.0, 0),
                      (d3, "AA-3", None, 0, 12000.0)]},
            # Same reference across two rows — ONE invoice.
            {"name": f"Beta Traders {tag}", "tab": f"BETA {tag}", "opening": 0.0,
             "rows": [(d1, None, "AA-10", 0, 40000.0),
                      (d1, None, "AA-10", 0, 15000.0),
                      (d2, None, "BAH # 11841307", 20000.0, 0)]},
            # Negative opening — customer paid ahead.
            {"name": f"Gamma Supply {tag}", "tab": f"GAMMA {tag}", "opening": -75000.0,
             "rows": [(d2, None, "AA-20", 0, 30000.0)]},
            # Sheet name differs from the index name — must be offered, not assumed.
            {"name": f"Delta Developers {tag}", "tab": f"DELTA {tag}", "opening": 5000.0,
             "sheetName": f"Delta Developers & Builders (PVT) LTD {tag}",
             "rows": [(d3, None, "AA-30", 0, 9000.0)]},
        ]
        book = ledger_workbook(clients_spec)

        r = upload(ledger_preview, h, book, "ledger.xlsx", lform, lparams)
        lp = r.json() if r.ok else {}
        check("the ledger workbook previews", r.ok, f"http {r.status_code}: {r.text[:180]}")
        check("every customer on the index is read", len(lp.get("clients", [])) == 4,
              f"{len(lp.get('clients', []))} customers")
        check("every customer reconciles against the index",
              lp.get("clientsOutOfBalance") == 0 and lp.get("canCommit") is True,
              f"outOfBalance={lp.get('clientsOutOfBalance')} errors={lp.get('blockingErrors')}")

        byname = {c["indexName"]: c for c in lp.get("clients", [])}
        alpha = byname.get(f"Alpha Hardware {tag}")
        check("closing is opening plus credit minus debit",
              alpha is not None and abs(alpha["computedClosing"] - 157000.0) < 0.01,
              f"closing={alpha['computedClosing'] if alpha else '-'}")
        check("an undated row is dated from the row above",
              alpha is not None and alpha["undatedRowCount"] == 1,
              f"undated={alpha['undatedRowCount'] if alpha else '-'}")

        beta = byname.get(f"Beta Traders {tag}")
        beta_invoices = [i for i in lp["invoices"] if i["indexRow"] == beta["indexRow"]]
        check("one reference across two rows becomes one invoice",
              len(beta_invoices) == 1 and abs(beta_invoices[0]["amount"] - 55000.0) < 0.01,
              f"{len(beta_invoices)} invoices, amount={beta_invoices[0]['amount'] if beta_invoices else '-'}")
        check("the merge is reported, not silent",
              any("appears on 2 rows" in w for w in beta["warnings"]),
              f"warnings={beta['warnings']}")

        beta_receipt = [x for x in lp["receipts"] if x["indexRow"] == beta["indexRow"]][0]
        check("a bank reference becomes a bank transfer, keeping the reference",
              beta_receipt["method"] == "Bank Transfer" and "11841307" in (beta_receipt["description"] or ""),
              f"method={beta_receipt['method']} desc={beta_receipt['description']}")

        alpha_receipt = [x for x in lp["receipts"] if x["indexRow"] == alpha["indexRow"]][0]
        check("a cash row becomes a cash receipt", alpha_receipt["method"] == "Cash",
              f"method={alpha_receipt['method']}")

        gamma = byname.get(f"Gamma Supply {tag}")
        gamma_open_receipts = [x for x in lp["receipts"]
                               if x["indexRow"] == gamma["indexRow"] and x["isOpening"]]
        gamma_open_invoices = [i for i in lp["invoices"]
                               if i["indexRow"] == gamma["indexRow"] and i["isOpening"]]
        check("a negative opening becomes a receipt, never a negative invoice",
              len(gamma_open_receipts) == 1 and len(gamma_open_invoices) == 0
              and abs(gamma_open_receipts[0]["amount"] - 75000.0) < 0.01,
              f"receipts={len(gamma_open_receipts)} invoices={len(gamma_open_invoices)}")

        delta = byname.get(f"Delta Developers {tag}")
        check("a differing sheet name is surfaced for confirmation",
              delta is not None and delta["sheetName"] is not None,
              f"sheetName={delta['sheetName'] if delta else '-'}")

        # Two positive openings (Alpha, Delta). Beta opens at zero, and Gamma's
        # negative opening is a receipt, so neither produces an opening invoice.
        openings = [i for i in lp["invoices"] if i["isOpening"]]
        check("opening invoices are numbered clear of the sheet's own references",
              all(i["invoiceNumber"] >= 900000 for i in openings) and len(openings) == 2,
              f"{len(openings)} openings, numbers={[i['invoiceNumber'] for i in openings]}")

        before = requests.get(f"{api}/clients/company/{other}", headers=h, timeout=60).json()
        check("ledger preview wrote nothing", len(before) == 0, f"{len(before)} clients exist")

        # ── 10. Ledger commit ───────────────────────────────────────────
        lbody = {
            "companyId": other,
            "fileSha256": lp["fileSha256"], "fileName": "ledger.xlsx",
            "fileSizeBytes": lp["fileSizeBytes"],
            "openingDate": lp["openingDate"], "periodEnd": lp["periodEnd"],
            "setGlCutover": True,
            "clients": lp["clients"], "invoices": lp["invoices"], "receipts": lp["receipts"],
        }
        requests.post(f"{api}/accounts/company/{other}/seed-wholesale", headers=h, timeout=120)
        r = requests.post(ledger_commit, headers=h, timeout=600, json=lbody)
        lres = r.json() if r.ok else {}
        check("ledger commit succeeds", r.ok, f"http {r.status_code}: {r.text[:200]}")
        check("four customers are created", lres.get("clientsCreated") == 4,
              f"created={lres.get('clientsCreated')}")
        # Six from the sheets (AA-10's two rows are one invoice) plus the two openings.
        check("eight invoices are created", lres.get("invoicesCreated") == 8,
              f"invoices={lres.get('invoicesCreated')}")
        check("three receipts are created", lres.get("receiptsCreated") == 3,
              f"receipts={lres.get('receiptsCreated')}")

        expected_total = sum(c["opening"] + sum(r_[4] for r_ in c["rows"]) - sum(r_[3] for r_ in c["rows"])
                             for c in clients_spec)
        check("the receivable total matches the workbook",
              abs(lres.get("totalReceivable", 0) - expected_total) < 0.05,
              f"receivable={lres.get('totalReceivable')} expected={expected_total}")
        check("the ledger is frozen at the period end",
              (lres.get("glLockDate") or "").startswith("2026-06-30"),
              f"lock={lres.get('glLockDate')}")

        # Read the balances back through the app's OWN ledger, not the importer.
        created = requests.get(f"{api}/clients/company/{other}", headers=h, timeout=60).json()
        ids = {c["name"]: c["id"] for c in created}
        total_back = 0.0
        drift = []
        for c in lp["clients"]:
            cname = c["sheetName"] or c["indexName"]
            cid_ = ids.get(cname)
            if cid_ is None:
                drift.append((cname, "client not found"))
                continue
            L = requests.get(f"{api}/customer-ledger/company/{other}/client/{cid_}", headers=h, timeout=60)
            if not L.ok:
                drift.append((cname, f"http {L.status_code}"))
                continue
            closing = L.json()["closingBalance"]
            total_back += closing
            if abs(closing - c["computedClosing"]) > 0.01:
                drift.append((cname, f"{closing} vs {c['computedClosing']}"))

        check("every customer's ledger reads back the previewed balance", not drift,
              f"{len(drift)} differ: {drift[:3]}")
        check("the ledger total matches the workbook",
              abs(total_back - expected_total) < 0.05,
              f"ledger total={total_back} expected={expected_total}")

        r = requests.get(f"{api}/accounts/company/{other}/flat", headers=h, timeout=60)
        ar = next((a for a in r.json() if a.get("controlType") == "AccountsReceivable"), None) if r.ok else None
        check("Accounts receivable carries the opening balance",
              ar is not None and abs(ar.get("openingBalance", 0) - expected_total) < 0.05
              and ar.get("openingBalanceIsDebit") is True,
              f"opening={ar.get('openingBalance') if ar else '-'}")

        check("a customer in advance shows a credit balance",
              abs(requests.get(f"{api}/customer-ledger/company/{other}/client/{ids[f'Gamma Supply {tag}']}",
                               headers=h, timeout=60).json()["closingBalance"] - (-45000.0)) < 0.01)

        # ── 11. Ledger re-import and number collisions ──────────────────
        r = upload(ledger_preview, h, book, "ledger.xlsx", lform, lparams)
        lp2 = r.json() if r.ok else {}
        check("re-uploading the ledger is refused",
              any("already imported" in e for e in lp2.get("blockingErrors", [])),
              f"errors={lp2.get('blockingErrors')}")
        # Imported documents are numbered from the reserved band and keep the
        # workbook's reference, so the second defence is about the REFERENCES
        # already present, not about invoice numbers -- re-importing would
        # double those customers' balances rather than collide on a number.
        check("the documents already imported are named, by their own reference",
              any("already imported into this company" in e for e in lp2.get("blockingErrors", [])),
              f"errors={lp2.get('blockingErrors')}")
        check("imported documents never take the company's invoice numbers",
              all(int(i["invoiceNumber"]) >= 900001
                  for i in requests.get(f"{api}/invoices/company/{other}",
                                        headers=h, timeout=120).json()),
              "an imported document landed in the operator's sequence")

        # The operator's own bill must be the FIRST row, not buried behind the
        # migrated history. Imported documents are numbered from the reserved
        # 900001+ band, so a list ordered by invoice number put every one of
        # them ahead of a bill raised a minute ago -- on a real import that is
        # hundreds of rows, i.e. the operator's bill on the last page.
        cl = requests.get(f"{api}/clients/company/{other}", headers=h, timeout=60).json()
        it = requests.get(f"{api}/itemtypes", headers=h, timeout=60).json()
        if cl and it:
            r = requests.post(f"{api}/invoices/standalone", headers=h, timeout=120, json={
                # Dated TODAY, not in the workbook's period: the ledger commit
                # sets Company.GlLockDate, so a bill back-dated into the
                # imported period is refused -- correctly.
                "date": datetime.now().strftime("%Y-%m-%d"), "companyId": other, "clientId": cl[0]["id"], "gstRate": 0,
                "items": [{"description": f"AFTER IMPORT {tag}", "quantity": 1,
                           "uom": "Pcs", "unitPrice": 100, "itemTypeId": it[0]["id"]}],
            })
            fresh = r.json() if r.ok else {}
            check("a bill raised after the import is accepted", r.ok, f"{r.status_code} {r.text[:200]}")
            page1 = requests.get(f"{api}/invoices/company/{other}/paged",
                                 headers=h, timeout=120,
                                 params={"page": 1, "pageSize": 10}).json()
            rows = page1.get("items") or page1.get("data") or []
            check("the newest bill is the first row on page 1",
                  bool(rows) and rows[0].get("id") == fresh.get("id"),
                  f"first row = {rows[0] if rows else None}")
            check("it is not ordered by invoice number",
                  bool(rows) and int(rows[0].get("invoiceNumber") or 0) < 900001,
                  f"first row number = {rows[0].get('invoiceNumber') if rows else None}")
        else:
            skip("the newest bill is the first row on page 1", "no client or item type available")

        clash = ledger_workbook([
            {"name": f"One {tag}", "tab": f"ONE {tag}", "opening": 0.0,
             "rows": [(d1, None, "ZZ-77", 0, 1000.0)]},
            {"name": f"Two {tag}", "tab": f"TWO {tag}", "opening": 0.0,
             "rows": [(d1, None, "ZZ-77", 0, 2000.0)]},
        ])
        r = upload(ledger_preview, h, clash, "clash.xlsx", lform, {"companyId": company})
        lp3 = r.json() if r.ok else {}
        check("one reference used by two customers is refused",
              any("used by more than one customer" in e for e in lp3.get("blockingErrors", [])),
              f"errors={lp3.get('blockingErrors')}")

        broken = ledger_workbook([
            {"name": f"Wrong {tag}", "tab": f"WRONG {tag}", "opening": 0.0,
             "rows": [(d1, None, "QQ-1", 0, 1000.0)]}])
        wb_ = openpyxl.load_workbook(io.BytesIO(broken))
        wb_["Index"].cell(3, 6, 999999.0)          # index states a closing the rows cannot produce
        buf = io.BytesIO(); wb_.save(buf)
        r = upload(ledger_preview, h, buf.getvalue(), "broken.xlsx", lform, {"companyId": company})
        lp4 = r.json() if r.ok else {}
        check("a customer that does not reconcile blocks the import",
              lp4.get("clientsOutOfBalance") == 1 and lp4.get("canCommit") is False,
              f"outOfBalance={lp4.get('clientsOutOfBalance')} canCommit={lp4.get('canCommit')}")

        # ── 12. Built-in layouts ────────────────────────────────────────
        # The product ships a layout per kind so a first import never starts
        # from a blank mapping form.
        r = requests.get(f"{api}/import-profiles", headers=h, timeout=30,
                         params={"companyId": company})
        allp = r.json() if r.ok else []
        builtins = [p_ for p_ in allp if p_.get("isDefault")]
        check("a built-in layout ships for each kind",
              {p_["kind"] for p_ in builtins} == {"OpeningStock", "CustomerLedger"},
              f"found {[(p_['kind'], p_['name']) for p_ in builtins]}")
        check("built-in layouts are installation-wide",
              all(p_["isShared"] for p_ in builtins),
              f"{[p_['name'] for p_ in builtins if not p_['isShared']]} are not shared")

        stock_builtin = next((p_ for p_ in builtins if p_["kind"] == "OpeningStock"), None)
        r = requests.delete(f"{api}/import-profiles/{stock_builtin['id']}", headers=h, timeout=30)
        check("a built-in layout cannot be deleted", r.status_code == 400,
              f"http {r.status_code}")

        # An unfamiliar workbook still gets a described layout to start from.
        odd = stock_workbook(base_rows(tag), month="Jul 2026", sheet_name="Nothing Like It")
        wbx = openpyxl.load_workbook(io.BytesIO(odd))
        wsx = wbx["Nothing Like It"]
        wsx.cell(1, 6, "Some Other Business Entirely")
        for col in range(1, 20):
            if wsx.cell(3, col).value:
                wsx.cell(3, col, f"Unfamiliar Heading {col}")
        bx = io.BytesIO(); wbx.save(bx)
        r = upload(idurl, h, bx.getvalue(), "odd.xlsx", None,
                   {"companyId": company, "kind": "OpeningStock"})
        odd_id = r.json() if r.ok else {}
        check("an unrecognised workbook is offered the built-in as a starting point",
              (odd_id.get("defaultProfile") or {}).get("isDefault") is True,
              f"default={odd_id.get('defaultProfile')}")
        check("and it is not passed off as a confident match",
              odd_id.get("matchedProfile") is None,
              f"matched={odd_id.get('matchedProfile')}")

        # ── 13. A layout survives its data changing ─────────────────────
        # The property the whole feature rests on: next period's workbook has
        # the same columns and entirely different contents, and must still be
        # recognised without being re-mapped.
        r = upload(idurl, h, stock_workbook(base_rows(tag)), "m1.xlsx", None,
                   {"companyId": company, "kind": "OpeningStock"})
        first = r.json().get("signatureHash") if r.ok else None

        later = [("ZZZZ-LOT-9", "7318", "7318.1510:-", f"ENTIRELY OTHER GOODS {tag}", "Fasteners", "Kg", 812.5, 990125),
                 ("ZZZZ-LOT-9", "9403", "9403.2000:-", f"SOMETHING ELSE AGAIN {tag}", "Furniture", "Pcs", 4, 61000)]
        r = upload(idurl, h, stock_workbook(later, month="Dec 2027"), "m2.xlsx", None,
                   {"companyId": company, "kind": "OpeningStock"})
        second = r.json().get("signatureHash") if r.ok else None
        check("the same layout with different data is still recognised",
              first is not None and first == second,
              f"{first} vs {second}")

        # ── 14. The period belongs to the import, not the layout ────────
        r = upload(ledger_preview, h, book, "ledger.xlsx",
                   {"mappingJson": json.dumps({k: v for k, v in LEDGER_MAPPING.items()
                                               if k not in ("periodStart", "periodEnd")})},
                   {"companyId": company})
        msg = (r.json().get("message") or "") if r.status_code == 400 else ""
        check("a layout with no period asks for one instead of guessing",
              r.status_code == 400 and "period" in msg.lower(),
              f"http {r.status_code}: {msg[:80]}")

        r = upload(ledger_preview, h, book, "ledger.xlsx",
                   {"mappingJson": json.dumps({k: v for k, v in LEDGER_MAPPING.items()
                                               if k not in ("periodStart", "periodEnd")}),
                    "periodStart": "2025-07-01", "periodEnd": "2026-06-30",
                    "openingDate": "2025-06-30"},
                   {"companyId": company})
        check("the period supplied with the import is used",
              r.ok and (r.json().get("periodEnd") or "").startswith("2026-06-30"),
              f"http {r.status_code}, periodEnd={r.json().get('periodEnd') if r.ok else '-'}")

        # ── 8. Isolation ────────────────────────────────────────────────
        r = upload(stock_preview, h, good, "stock.xlsx", form, {"companyId": 999999})
        check("preview refuses a company that does not exist", r.status_code == 404,
              f"http {r.status_code}")

        r = requests.post(stock_commit, headers=h, timeout=60,
                          json=dict(body, companyId=999999))
        check("commit refuses a company that does not exist", r.status_code == 404,
              f"http {r.status_code}")

        r = upload(stock_preview, h, good, "stock.xlsx",
                   {"mappingJson": json.dumps(STOCK_MAPPING)}, {"companyId": other},
                   )
        check("a layout is not required to belong to the company being imported into",
              r.ok, f"http {r.status_code}")

        r = upload(stock_preview, h, good, "stock.xlsx", None,
                   {"companyId": other, "profileId": pid})
        check("another company's private layout cannot be used", r.status_code == 404,
              f"http {r.status_code}")

        r = upload(ledger_preview, h, book, "ledger.xlsx", lform, {"companyId": 999999})
        check("ledger preview refuses a company that does not exist", r.status_code == 404,
              f"http {r.status_code}")

        r = requests.post(ledger_commit, headers=h, timeout=60, json=dict(lbody, companyId=999999))
        check("ledger commit refuses a company that does not exist", r.status_code == 404,
              f"http {r.status_code}")

    finally:
        if not args.keep:
            for p in profile_ids:
                requests.delete(f"{api}/import-profiles/{p}", headers=h, timeout=60)
            for c in (company, other):
                requests.delete(f"{api}/companies/{c}", headers=h, timeout=300)

    return report()


def report():
    failed = [r for r in results if r[0] == FAIL]
    skipped = [r for r in results if r[0] == SKIP]
    passed = [r for r in results if r[0] == PASS]
    print(f"\n{len(passed)} passed, {len(failed)} failed, {len(skipped)} skipped")
    if failed:
        print("FAILURES:")
        for _, name, detail in failed:
            print(f"  - {name}: {detail}")
        return 1
    print("all PASS")
    return 0


if __name__ == "__main__":
    sys.exit(main())
