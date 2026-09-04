"""
Live end-to-end test for the Stock dashboard Excel export.

  GET /api/stock/company/{id}/onhand/excel?search=

The offline harness (scripts/stock_export_harness) already pins the LAYOUT
against synthetic rows. This suite pins the things only a running server can
answer:

  1. the workbook's figures equal what GET .../onhand returns, row for row —
     the export and the grid come out of one valuation walk, and the whole
     point of that is that they cannot disagree;
  2. the totals row ties to the item rows AND to the API's own sum;
  3. the drill-down agrees with the movements feed;
  4. the two-permission split behaves: stock.dashboard.export alone yields a
     workbook that SAYS its movement detail is missing, both permissions yield
     the collapsed drill-down, and neither yields 403;
  5. the search term reaches the workbook and is named on its face.

Any user or role this suite creates is deleted in a finally block — the local
branch database is meant to stay at one company.

Usage:
  python scripts/test_stock_export_excel.py                        # localhost:5134, first company
  python scripts/test_stock_export_excel.py --base http://localhost:5136 --company 451

Exit code 0 = every assertion passed.
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import tempfile
import urllib.error
import urllib.parse
import urllib.request
from decimal import Decimal

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

BASE = "http://localhost:5134"
OUT = tempfile.mkdtemp(prefix="stock-export-")

# Column layout — must match Helpers/StockExcelBuilder.cs.
C_ITEM, C_HS, C_UOM = 1, 2, 3
C_OPEN, C_IN, C_OUT, C_ONHAND = 4, 5, 6, 7
C_UNIT, C_EXCL, C_RATE, C_TAX, C_INCL = 8, 9, 10, 11, 12

# DTO field -> column, for the row-for-row comparison against the grid.
FIELD_COLUMNS = {
    "openingBalance": C_OPEN,
    "totalIn": C_IN,
    "totalOut": C_OUT,
    "onHand": C_ONHAND,
    "unitCost": C_UNIT,
    "valueExcludingTax": C_EXCL,
    "salesTaxRate": C_RATE,
    "salesTax": C_TAX,
    "valueIncludingTax": C_INCL,
}

passed = 0
failed = 0


def check(suite: str, name: str, ok: bool, detail: str = "") -> None:
    global passed, failed
    if ok:
        passed += 1
        print(f"  [PASS] {name}")
    else:
        failed += 1
        print(f"  [FAIL] {name}" + (f"   — {detail}" if detail else ""))


def request(method: str, path: str, token: str | None = None,
            body=None, binary: bool = False):
    data = json.dumps(body).encode("utf-8") if body is not None else None
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(BASE + path, data=data, method=method, headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=120) as r:
            blob = r.read()
            if binary:
                return r.status, blob, dict(r.headers)
            text = blob.decode("utf-8") if blob else ""
            return r.status, (json.loads(text) if text else None), dict(r.headers)
    except urllib.error.HTTPError as e:
        blob = e.read() if e.fp else b""
        try:
            return e.code, json.loads(blob.decode("utf-8")), dict(e.headers or {})
        except Exception:
            return e.code, blob[:300], dict(e.headers or {})


def login(username: str, password: str) -> str:
    status, data, _ = request("POST", "/api/auth/login",
                              body={"username": username, "password": password})
    if status != 200:
        sys.exit(f"login failed for {username}: {status} {data}")
    return data["token"]


def save(blob: bytes, name: str) -> str:
    path = os.path.join(OUT, name)
    with open(path, "wb") as fh:
        fh.write(blob)
    return path


def dec(value) -> Decimal:
    return Decimal(str(value if value is not None else 0))


def sheet_of(path: str):
    return openpyxl.load_workbook(path).active


def outline_level(ws, row: int) -> int:
    dim = ws.row_dimensions.get(row)
    return dim.outlineLevel if dim else 0


def anatomy(ws):
    """(header row, item rows, grouped rows, total row) of a stock workbook."""
    header = next(r for r in range(1, 40) if ws.cell(r, C_ITEM).value == "Item")
    total = next(r for r in range(header, ws.max_row + 1)
                 if str(ws.cell(r, C_ITEM).value or "").startswith("TOTAL"))
    items = [r for r in range(header + 1, total) if outline_level(ws, r) == 0]
    grouped = [r for r in range(header + 1, total) if outline_level(ws, r) > 0]
    return header, items, grouped, total


def column_text(ws, row: int) -> str:
    return " ".join(str(ws.cell(row, c).value or "") for c in range(1, 15))


def all_text(ws) -> str:
    return " ".join(column_text(ws, r) for r in range(1, ws.max_row + 1))


def rendered(cell) -> str:
    """What Excel will actually paint in the cell — the thing that can clip."""
    value = cell.value
    if value is None:
        return ""
    if hasattr(value, "year"):
        return "00-00-0000"
    if isinstance(value, (int, float)):
        fmt = cell.number_format or ""
        if "0.####" in fmt:
            text = f"{abs(value):,.4f}".rstrip("0").rstrip(".")
        else:
            text = f"{abs(value):,.4f}" if "0.0000" in fmt else f"{abs(value):,.2f}"
        if value < 0:
            text = "-" + text
        if '"%"' in fmt:
            text += "%"
        return text
    return str(value)


def clipped_cells(ws) -> list[str]:
    widths = {letter: dim.width for letter, dim in ws.column_dimensions.items()}
    merged = {c.coordinate for rng in ws.merged_cells.ranges
              for row in ws[str(rng)] for c in row}
    out = []
    for row in ws.iter_rows():
        for cell in row:
            # A merged banner spans all 14 columns, and a wrapped cell grows its
            # row instead of clipping — neither can lose text.
            if cell.coordinate in merged:
                continue
            if cell.alignment and cell.alignment.wrap_text:
                continue
            text = rendered(cell)
            if not text:
                continue
            width = widths.get(cell.column_letter)
            if width is None:
                continue
            indent = (cell.alignment.indent or 0) if cell.alignment else 0
            if len(text) + indent > width:
                out.append(f"{cell.coordinate} needs {len(text) + indent:.0f} "
                           f"has {width:.1f}: {text[:40]}")
    return out


def main() -> int:
    global BASE

    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--base", default=BASE)
    ap.add_argument("--company", type=int, default=None,
                    help="company id to export (default: the first accessible one)")
    args = ap.parse_args()
    BASE = args.base.rstrip("/")

    admin = login("admin", "admin123")

    status, companies, _ = request("GET", "/api/companies", token=admin)
    if status != 200 or not companies:
        sys.exit(f"could not list companies: {status} {companies}")
    cid = args.company or companies[0]["id"]
    company = next((c for c in companies if c["id"] == cid), None)
    if company is None:
        sys.exit(f"company {cid} is not accessible to admin")
    expected_banner = company.get("brandName") or company["name"]
    print(f"\nBase {BASE}  ·  company {cid} ({expected_banner})")

    made_users: list[int] = []
    made_roles: list[int] = []

    def provision(username: str, role_name: str, keys: list[str]) -> str:
        """A throwaway user holding exactly `keys`, with access to `cid`."""
        # Clear leftovers from an interrupted run — the names are unique.
        _, users, _ = request("GET", "/api/users", token=admin)
        for u in users if isinstance(users, list) else []:
            if u["username"] == username:
                request("DELETE", f"/api/users/{u['id']}", token=admin)
        _, roles, _ = request("GET", "/api/roles", token=admin)
        for r in roles if isinstance(roles, list) else []:
            if r["name"] == role_name:
                request("DELETE", f"/api/roles/{r['id']}", token=admin)

        st, role, _ = request("POST", "/api/roles", token=admin, body={
            "name": role_name,
            "description": "stock export permission probe (test)",
            "permissionKeys": keys,
        })
        assert st in (200, 201), f"create role {role_name}: {st} {role}"
        made_roles.append(role["id"])

        st, user, _ = request("POST", "/api/users", token=admin, body={
            "username": username, "password": "test1234", "fullName": username,
            "role": role_name, "roleIds": [role["id"]], "companyIds": [cid],
        })
        assert st in (200, 201), f"create user {username}: {st} {user}"
        made_users.append(user["id"])
        return login(username, "test1234")

    try:
        # ── Suite 1: the workbook cannot disagree with the grid ──────────────
        print("\n  Suite 1 — the workbook equals the on-hand grid")
        _, grid, _ = request("GET", f"/api/stock/company/{cid}/onhand", token=admin)
        status, blob, headers = request(
            "GET", f"/api/stock/company/{cid}/onhand/excel", token=admin, binary=True)
        check("s1", "export returns 200", status == 200, f"got {status}")
        if status != 200:
            return 1
        check("s1", "served as an .xlsx attachment",
              "spreadsheetml" in headers.get("Content-Type", "")
              and ".xlsx" in headers.get("Content-Disposition", ""),
              f"{headers.get('Content-Type')} / {headers.get('Content-Disposition')}")

        path = save(blob, "onhand.xlsx")
        ws = sheet_of(path)
        header, items, grouped, total = anatomy(ws)
        print(f"    {len(grid)} items on the grid; workbook has {len(items)} item rows, "
              f"{len(grouped)} grouped rows")

        check("s1", "one workbook row per on-hand item", len(items) == len(grid),
              f"{len(items)} vs {len(grid)}")
        check("s1", "banner carries the company name (brand wins)",
              ws.cell(1, C_ITEM).value == expected_banner,
              f"{ws.cell(1, C_ITEM).value!r} vs {expected_banner!r}")

        by_name = {r["itemTypeName"]: r for r in grid}
        drift = []
        for row in items:
            name = ws.cell(row, C_ITEM).value
            source = by_name.get(name)
            if source is None:
                drift.append(f"r{row}: {name!r} absent from the grid")
                continue
            for field, col in FIELD_COLUMNS.items():
                want, got = dec(source[field]), dec(ws.cell(row, col).value)
                if want != got:
                    drift.append(f"r{row} {name!r}.{field}: sheet {got} vs grid {want}")
        check("s1", "every figure matches the grid exactly", not drift,
              " | ".join(drift[:4]))

        # ── Suite 2: totals ─────────────────────────────────────────────────
        print("\n  Suite 2 — totals tie to the rows and to the API")
        for field, col in [("openingBalance", C_OPEN), ("totalIn", C_IN),
                           ("totalOut", C_OUT), ("onHand", C_ONHAND),
                           ("valueExcludingTax", C_EXCL), ("salesTax", C_TAX),
                           ("valueIncludingTax", C_INCL)]:
            summed = sum(dec(ws.cell(r, col).value) for r in items)
            stated = dec(ws.cell(total, col).value)
            api_sum = sum(dec(r[field]) for r in grid)
            check("s2", f"TOTAL {field} = rows = API",
                  summed == stated == api_sum,
                  f"rows {summed} / total {stated} / api {api_sum}")
        # A weighted average and a percentage do not add up.
        check("s2", "TOTAL leaves Unit Cost blank", ws.cell(total, C_UNIT).value is None)
        check("s2", "TOTAL leaves Tax Rate blank", ws.cell(total, C_RATE).value is None)

        # ── Suite 3: the drill-down ─────────────────────────────────────────
        print("\n  Suite 3 — the drill-down agrees with the movements feed")
        _, feed, _ = request("GET",
                             f"/api/stock/company/{cid}/movements?pageSize=200",
                             token=admin)
        rows = feed["items"]
        print(f"    movements feed: {feed['totalCount']} rows")
        if not rows:
            print("    [skip] no movements on this company — nothing to nest")
        else:
            check("s3", "workbook nests movement rows", len(grouped) > 0)
            check("s3", "every nested row starts hidden (collapsed)",
                  all(ws.row_dimensions[r].hidden for r in grouped),
                  f"{sum(1 for r in grouped if not ws.row_dimensions[r].hidden)} visible")
            # One sub-header per grouped item, then that item's movement lines.
            sub_headers = [r for r in grouped if ws.cell(r, C_ITEM).value == "Date"]
            lines = [r for r in grouped if r not in sub_headers]
            check("s3", "one column sub-header per grouped item",
                  0 < len(sub_headers) <= len(items), f"{len(sub_headers)} sub-headers")
            # Same-document folding can only ever REDUCE the line count.
            check("s3", "movement lines never exceed the feed",
                  0 < len(lines) <= min(feed["totalCount"], 200),
                  f"{len(lines)} lines vs {feed['totalCount']} feed rows")
            sheet_qty = sum(dec(ws.cell(r, C_IN).value) + dec(ws.cell(r, C_OUT).value)
                            for r in lines)
            feed_qty = sum(dec(m["quantity"]) for m in rows)
            check("s3", "nested quantities sum to the feed's quantities",
                  sheet_qty == feed_qty, f"sheet {sheet_qty} vs feed {feed_qty}")
            check("s3", "a movement never fills both Qty In and Qty Out",
                  all(ws.cell(r, C_IN).value is None or ws.cell(r, C_OUT).value is None
                      for r in lines))

        # ── Suite 4: nothing clipped, on real data ──────────────────────────
        print("\n  Suite 4 — nothing is cut off")
        clipped = clipped_cells(ws)
        check("s4", "no clipped cell in the full export", not clipped,
              " | ".join(clipped[:4]))

        # ── Suite 5: the two-permission split ───────────────────────────────
        print("\n  Suite 5 — export vs movement-detail permissions")
        tok = provision("stkexp_nomov", "StkExport NoMovements (test)",
                        ["stock.dashboard.view", "stock.dashboard.export"])
        status, blob, _ = request("GET", f"/api/stock/company/{cid}/onhand/excel",
                                  token=tok, binary=True)
        check("s5", "stock.dashboard.export alone returns 200", status == 200, f"got {status}")
        if status == 200:
            ws_nm = sheet_of(save(blob, "no-movements.xlsx"))
            check("s5", "that workbook carries NO drill-down",
                  all(outline_level(ws_nm, r) == 0 for r in range(1, ws_nm.max_row + 1)))
            text = all_text(ws_nm)
            check("s5", "it says movement detail is not included",
                  "Movement detail is not included" in text)
            check("s5", "its provenance line admits the omission",
                  "Movement detail omitted" in text)
            check("s5", "it still carries the figures and a total",
                  "TOTAL" in text and ws_nm.cell(6, C_ONHAND).value == "On Hand")
            check("s5", "no clipped cell without the drill-down",
                  not clipped_cells(ws_nm))

        tok = provision("stkexp_full", "StkExport Full (test)",
                        ["stock.dashboard.view", "stock.dashboard.export",
                         "stock.movements.view"])
        status, blob, _ = request("GET", f"/api/stock/company/{cid}/onhand/excel",
                                  token=tok, binary=True)
        check("s5", "both permissions return 200", status == 200, f"got {status}")
        if status == 200 and rows:
            ws_full = sheet_of(save(blob, "with-movements.xlsx"))
            g = [r for r in range(1, ws_full.max_row + 1) if outline_level(ws_full, r) > 0]
            check("s5", "that workbook carries the drill-down", len(g) > 0)
            check("s5", "and it starts collapsed",
                  all(ws_full.row_dimensions[r].hidden for r in g))
            check("s5", "its legend explains how to expand",
                  "start collapsed" in all_text(ws_full))

        tok = provision("stkexp_none", "StkExport ViewOnly (test)",
                        ["stock.dashboard.view"])
        status, body, _ = request("GET", f"/api/stock/company/{cid}/onhand/excel", token=tok)
        check("s5", "403 without stock.dashboard.export", status == 403, f"got {status} {body}")
        status, _, _ = request("GET", f"/api/stock/company/{cid}/onhand", token=tok)
        check("s5", "the grid itself still works for that user", status == 200, f"got {status}")

        # ── Suite 6: the search term ────────────────────────────────────────
        print("\n  Suite 6 — the search reaches the workbook")
        term = grid[0]["itemTypeName"][:6]
        expected = [r for r in grid
                    if term.lower() in r["itemTypeName"].lower()
                    or term.lower() in (r["hsCode"] or "").lower()]
        status, blob, _ = request(
            "GET", f"/api/stock/company/{cid}/onhand/excel"
                   f"?search={urllib.parse.quote(term)}", token=admin, binary=True)
        check("s6", "export with a search term returns 200", status == 200, f"got {status}")
        if status == 200:
            ws_s = sheet_of(save(blob, "searched.xlsx"))
            _, s_items, _, s_total = anatomy(ws_s)
            check("s6", f"narrowed to the {len(expected)} matching item(s)",
                  len(s_items) == len(expected),
                  f"{len(s_items)} rows vs {len(expected)} expected")
            check("s6", "the search is named on the sheet",
                  f'Search: "{term}"' in column_text(ws_s, 3), column_text(ws_s, 3)[:110])
            check("s6", "the narrowed total ties to the narrowed rows",
                  sum(dec(ws_s.cell(r, C_INCL).value) for r in s_items)
                  == dec(ws_s.cell(s_total, C_INCL).value))

    finally:
        print("\n  Cleanup")
        for uid in made_users:
            st, _, _ = request("DELETE", f"/api/users/{uid}", token=admin)
            print(f"    user {uid}: {st}")
        for rid in made_roles:
            st, _, _ = request("DELETE", f"/api/roles/{rid}", token=admin)
            print(f"    role {rid}: {st}")

    print(f"\n=== {passed}/{passed + failed} checks passed ===")
    if failed:
        print(f"{failed} FAILING CHECK(S). Workbooks kept in {OUT}")
        return 1
    print("STOCK EXPORT LIVE SUITE PASSED")
    return 0


if __name__ == "__main__":
    sys.exit(main())
