"""
Dump Excel file structure: cell values, merged cells, row heights, column widths.
Uses openpyxl for .xlsx and xlrd for .xls files.
"""
import os
import sys
import io

# Force UTF-8 output to handle special characters
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

FILES = [
    (r"C:\Users\hussahuz\Downloads\DC # 4161 Afroz Textile.xls", "Hakimi DC - xls"),
    (r"C:\Users\hussahuz\Downloads\DC # 1073 MEKO DENIM.xls", "Roshan DC - xls"),
    (r"C:\Users\hussahuz\Downloads\Bill # 3719 AFROZE TEXTILE.xlsx", "Hakimi Bill - xlsx"),
    (r"C:\Users\hussahuz\Downloads\Bill # 970 MEKO DENIM.xls", "Roshan Bill - xls"),
    (r"C:\Users\hussahuz\Downloads\INVOICE # 3719 AFROZE TEXTILE.xlsx", "Hakimi Tax Invoice - xlsx"),
    (r"C:\Users\hussahuz\Downloads\Invoice # 970 Meko DENIM.xlsx", "Roshan Tax Invoice - xlsx"),
]

def dump_xls(filepath, label):
    import xlrd
    print(f"\n{'='*100}")
    print(f"FILE: {label}")
    print(f"PATH: {filepath}")
    print(f"{'='*100}")

    if not os.path.exists(filepath):
        print(f"  *** FILE NOT FOUND ***")
        return

    wb = xlrd.open_workbook(filepath, formatting_info=True)

    for sheet_idx in range(wb.nsheets):
        sh = wb.sheet_by_index(sheet_idx)
        print(f"\n--- Sheet: '{sh.name}' | Rows: {sh.nrows} | Cols: {sh.ncols} ---")

        # Merged cells
        merged = sh.merged_cells  # list of (row_lo, row_hi, col_lo, col_hi)
        if merged:
            print(f"\n  MERGED CELLS ({len(merged)} ranges):")
            for rlo, rhi, clo, chi in merged:
                print(f"    rows {rlo}-{rhi-1}, cols {clo}-{chi-1}  (xlrd: {rlo}:{rhi}, {clo}:{chi})")
        else:
            print(f"\n  MERGED CELLS: none")

        # Column widths
        print(f"\n  COLUMN WIDTHS:")
        for col_idx in range(sh.ncols):
            # xlrd stores column widths in the colinfo dict
            if col_idx in sh.colinfo_map:
                ci = sh.colinfo_map[col_idx]
                # width is in 1/256th of the width of '0' character
                width_chars = ci.width / 256.0
                print(f"    Col {col_idx}: {ci.width} units ({width_chars:.2f} chars)")
            else:
                print(f"    Col {col_idx}: default")

        # Row heights
        print(f"\n  ROW HEIGHTS (explicit only):")
        found_heights = False
        for row_idx in range(sh.nrows):
            if row_idx in sh.rowinfo_map:
                ri = sh.rowinfo_map[row_idx]
                # height is in twips (1/20 of a point)
                height_pt = ri.height / 20.0
                print(f"    Row {row_idx}: {ri.height} twips ({height_pt:.1f} pt)")
                found_heights = True
        if not found_heights:
            print(f"    (none)")

        # Cell values
        print(f"\n  CELL VALUES:")
        for row_idx in range(sh.nrows):
            for col_idx in range(sh.ncols):
                cell = sh.cell(row_idx, col_idx)
                if cell.ctype != xlrd.XL_CELL_EMPTY:
                    val = str(cell.value)
                    if len(val) > 80:
                        val = val[:80] + "..."
                    ctype_name = {
                        0: "EMPTY", 1: "TEXT", 2: "NUMBER",
                        3: "DATE", 4: "BOOLEAN", 5: "ERROR", 6: "BLANK"
                    }.get(cell.ctype, str(cell.ctype))
                    print(f"    R{row_idx} C{col_idx}: [{ctype_name}] {val}")


def dump_xlsx(filepath, label):
    import openpyxl
    print(f"\n{'='*100}")
    print(f"FILE: {label}")
    print(f"PATH: {filepath}")
    print(f"{'='*100}")

    if not os.path.exists(filepath):
        print(f"  *** FILE NOT FOUND ***")
        return

    wb = openpyxl.load_workbook(filepath, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n--- Sheet: '{sheet_name}' | Rows: {ws.max_row} | Cols: {ws.max_column} ---")

        # Merged cells
        merged = list(ws.merged_cells.ranges)
        if merged:
            print(f"\n  MERGED CELLS ({len(merged)} ranges):")
            for mc in merged:
                print(f"    {mc}")
        else:
            print(f"\n  MERGED CELLS: none")

        # Column widths
        print(f"\n  COLUMN WIDTHS:")
        if ws.column_dimensions:
            for col_letter in sorted(ws.column_dimensions.keys()):
                cd = ws.column_dimensions[col_letter]
                w = cd.width
                if w is not None:
                    print(f"    Col {col_letter}: {w}")
                else:
                    print(f"    Col {col_letter}: default (None)")
        else:
            print(f"    (none)")

        # Row heights
        print(f"\n  ROW HEIGHTS (explicit only):")
        found_heights = False
        if ws.row_dimensions:
            for row_num in sorted(ws.row_dimensions.keys()):
                rd = ws.row_dimensions[row_num]
                if rd.height is not None:
                    print(f"    Row {row_num}: {rd.height} pt")
                    found_heights = True
        if not found_heights:
            print(f"    (none)")

        # Cell values
        print(f"\n  CELL VALUES:")
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if cell.value is not None:
                    val = str(cell.value)
                    if len(val) > 80:
                        val = val[:80] + "..."
                    print(f"    R{cell.row} C{cell.column} ({cell.coordinate}): {val}")


def main():
    for filepath, label in FILES:
        ext = os.path.splitext(filepath)[1].lower()
        if ext == '.xls':
            dump_xls(filepath, label)
        elif ext == '.xlsx':
            dump_xlsx(filepath, label)
        else:
            print(f"\n*** Unknown extension '{ext}' for {filepath}")

if __name__ == "__main__":
    main()
