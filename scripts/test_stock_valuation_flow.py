#!/usr/bin/env python3
"""
End-to-end stock VALUATION flow (Phase 2, 2026-09-02).

Proves that quantity and money move together, from the client's stock sheet all
the way through a purchase, a sale and both directions of adjustment:

    1. Import          the opening stock sheet WITH its value columns, and read
                       the five figures back off the stock dashboard's feed --
                       quantity, excluding, sales tax, rate, including.
    2. Purchase        an inward movement at a stated cost raises quantity AND
                       value, and re-averages the item.
    3. Invoice         an outward movement lowers quantity and takes value off
                       at the AVERAGE, never at the sale price. Selling at a
                       margin must not drain more value than the goods cost.
    4. Adjustments     up with a stated cost, up without one (valued at the
                       average), and down.
    5. Movements feed  every row carries its own unit cost, value and running
                       balance, and the last running value is the grid's value.
    6. Emptying        selling the last unit leaves value at exactly zero, with
                       no stray paisa left behind by rounding.
    7. Rate change     a purchase at a different rate re-prices the item, so
                       sales tax and the inclusive total follow it.

The workbook is synthetic by default so the suite runs anywhere. Pass
--stock-file to run the same flow against a real sheet through the SHIPPED
default layout (no mapping supplied), which is what an operator actually does:

    python scripts/test_stock_valuation_flow.py --base http://localhost:5135 \
        --stock-file "C:/path/Alpha Stock Sheet Aug 2026.xlsx"

Usage:
    python scripts/test_stock_valuation_flow.py [--base URL] [--keep]
"""

import argparse
import io
import json
import os
import sys
from datetime import date, datetime, timezone

try:
    import requests
except ImportError:
    print("requests is required:  pip install requests")
    sys.exit(2)

try:
    import openpyxl
except ImportError:
    print("openpyxl is required:  pip install openpyxl")
    sys.exit(2)

for _stream in (sys.stdout, sys.stderr):
    try:
        _stream.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, ValueError):
        pass

PASS, FAIL, SKIP = "PASS", "FAIL", "SKIP"
results = []
CENT = 0.005          # money is stored to 2dp; anything inside half a paisa agrees
QTY_TOL = 0.0001      # quantity is stored to 4dp


def check(name, ok, detail=""):
    results.append((PASS if ok else FAIL, name, detail))
    print(f"  [{PASS if ok else FAIL}] {name}" + (f" -- {detail}" if detail and not ok else ""))
    return ok


def skip(name, why):
    results.append((SKIP, name, why))
    print(f"  [{SKIP}] {name} -- {why}")


def near(a, b, tol=CENT):
    return abs(float(a or 0) - float(b or 0)) <= tol


# -- Workbook ---------------------------------------------------------------
# Same shape as the client's sheet: title band, headings on row 3, data from
# row 4, HS codes with a ":-" tail, and the Balance block at columns 18-21
# (qty, excluding, RATE, S.Tax). The rate is written as a FRACTION, as the real
# sheet writes it -- reading 0.18 as 18% is the whole point of the mapping.
STOCK_MAPPING = {
    "sheetSelect": {"mode": "byHeaderText", "mustContain": ["GD Number"]},
    "headerRow": 3,
    "firstDataRow": 4,
    "columns": {
        "lotRef": 2, "lotDate": 3, "hsCodeShort": 4, "hsCodeFull": 5,
        "itemName": 6, "unit": 9, "balanceQty": 18, "balanceValue": 19,
        "balanceTaxRate": 20, "balanceTax": 21,
    },
    "hsCodeStripSuffix": ":-",
    "ignoreColumns": [7],
}


def sheet_rows(tag):
    """Real Pakistan tariff lines -- validation is master-first, so an invented
    code is rejected. (lot, hs4, hs8, name, subcat, unit, qty, value, rate)"""
    return [
        ("KAPE-HC-5876", "8481", "8481.1000:-", f"SS BALL VALVE {tag}", "Valve", "Pcs", 12, 23160, 0.18),
        ("KAPE-HC-5876", "8513", "8513.1090:-", f"RECHARGEABLE LED {tag}", "Electrical", "Pcs", 46, 30000, 0.18),
        ("KAPE-HC-5876", "8450", "8450.9000:-", f"WASHING PARTS {tag}", "Electrical", "Kg", 67.5, 54160, 0.18),
        # Same item, second customs lot -- merges into one item type, and the
        # merged rate has to be value-weighted, not "whichever row came first".
        ("KAPW-HC-128753", "8450", "8450.9000:-", f"WASHING PARTS {tag}", "Electrical", "KG", 32.5, 26000, 0.18),
        ("KAPW-HC-128753", "8712", "8712.0000:-", f"CHILDREN BICYCLE {tag}", "Sports", "Pcs", 2, 15120, 0.25),
    ]


def stock_workbook(rows, month="Jul 2026"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = month
    ws.cell(1, 6, "ALPHA Trader")
    ws.cell(2, 1, "Stock Sheet")
    ws.cell(3, 1, month)
    headings = {1: "Claimed Month", 2: "GD Number", 3: "GD Date", 4: "4 Digit Hs Code",
                5: "8 Digit Hs code", 6: "Items", 7: "Sub Catory", 8: "Price",
                9: "Unit", 10: "Qty", 11: "Excluding", 12: "Rate", 13: "S.Tax",
                18: " Qty", 19: "Balance Exl", 20: "Rate", 21: "S.Tax"}
    for col, text in headings.items():
        ws.cell(3, col, text)

    for i, (lot, hs4, hs8, name, subcat, unit, qty, value, rate) in enumerate(rows):
        r = 4 + i
        ws.cell(r, 1, month)
        ws.cell(r, 2, lot)
        ws.cell(r, 3, "23-07-2025")
        ws.cell(r, 4, hs4)
        ws.cell(r, 5, hs8)
        ws.cell(r, 6, name)
        ws.cell(r, 7, subcat)
        ws.cell(r, 9, unit)
        ws.cell(r, 18, qty)
        ws.cell(r, 19, value)
        ws.cell(r, 20, rate)
        ws.cell(r, 21, round(value * rate, 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def sheet_totals(rows):
    """What the workbook itself says -- the figures the import must reproduce."""
    qty = sum(r[6] for r in rows)
    excl = sum(r[7] for r in rows)
    tax = sum(round(r[7] * r[8], 2) for r in rows)
    return qty, excl, tax, excl + tax


def read_real_sheet(path):
    """Totals straight out of a real workbook, so the assertions are the
    operator's own numbers rather than anything this script computed."""
    wb = openpyxl.load_workbook(path, data_only=True)
    target = None
    for name in wb.sheetnames:
        ws = wb[name]
        if any("GD Number" in str(ws.cell(3, c).value or "") for c in range(1, 26)):
            target = ws
            break
    if target is None:
        target = wb[wb.sheetnames[-1]]

    qty = excl = tax = 0.0
    for r in range(4, target.max_row + 1):
        name = target.cell(r, 6).value
        q = target.cell(r, 18).value
        if not name or q is None:
            continue                      # the totals row carries money but no item
        v = round(float(target.cell(r, 19).value or 0), 2)
        rate = float(target.cell(r, 20).value or 0)
        qty += float(q or 0)
        excl += v
        tax += round(v * rate, 2)
    return round(qty, 4), round(excl, 2), round(tax, 2), round(excl + tax, 2)


# -- HTTP -------------------------------------------------------------------
def api_call(method, url, h, **kw):
    kw.setdefault("timeout", 180)
    return requests.request(method, url, headers=h, **kw)


def upload(url, h, content, filename, data=None, params=None):
    return requests.post(
        url, params=params or {}, data=data or {},
        files={"file": (filename, io.BytesIO(content),
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=h, timeout=300)


def onhand_row(api, h, cid, item_id):
    r = api_call("GET", f"{api}/stock/company/{cid}/onhand", h)
    if not r.ok:
        return None
    return next((x for x in r.json() if x.get("itemTypeId") == item_id), None)


def onhand_totals(api, h, cid):
    r = api_call("GET", f"{api}/stock/company/{cid}/onhand", h)
    if not r.ok:
        return None
    rows = r.json()
    return {
        "qty": round(sum(float(x["onHand"]) for x in rows), 4),
        "excl": round(sum(float(x["valueExcludingTax"]) for x in rows), 2),
        "tax": round(sum(float(x["salesTax"]) for x in rows), 2),
        "incl": round(sum(float(x["valueIncludingTax"]) for x in rows), 2),
        "rows": rows,
    }


def movements(api, h, cid, item_id):
    """Every movement for one item, oldest first."""
    out, page = [], 1
    while True:
        r = api_call("GET", f"{api}/stock/company/{cid}/movements", h,
                     params={"itemTypeId": item_id, "page": page, "pageSize": 100})
        if not r.ok:
            break
        body = r.json()
        out += body.get("items", [])
        if len(out) >= body.get("totalCount", 0) or not body.get("items"):
            break
        page += 1
    return list(reversed(out))


# -- Suites -----------------------------------------------------------------
def suite_import(api, h, cid, workbook, mapping, expect, label):
    print(f"\n-- 1. Import the stock sheet with its values ({label}) --")
    params = {"companyId": cid}
    form = {"mappingJson": json.dumps(mapping)} if mapping else {}

    r = upload(f"{api}/spreadsheet-import/opening-stock/preview", h, workbook,
               "stock.xlsx", form, params)
    prev = r.json() if r.ok else {}
    if not check("the sheet previews", r.ok, f"http {r.status_code}: {r.text[:200]}"):
        return None

    e_qty, e_excl, e_tax, e_incl = expect
    check("preview quantity matches the sheet",
          near(prev.get("totalQuantity"), e_qty, QTY_TOL),
          f"got {prev.get('totalQuantity')} want {e_qty}")
    check("preview value excluding tax matches the sheet",
          near(prev.get("totalValue"), e_excl, 0.02),
          f"got {prev.get('totalValue')} want {e_excl}")
    check("preview sales tax matches the sheet",
          near(prev.get("totalSalesTax"), e_tax, 0.5),
          f"got {prev.get('totalSalesTax')} want {e_tax}")
    check("preview including tax is excluding + tax",
          near(prev.get("totalValueIncludingTax"),
               (prev.get("totalValue") or 0) + (prev.get("totalSalesTax") or 0)),
          f"got {prev.get('totalValueIncludingTax')}")
    check("the rate is read as a percentage, not a fraction",
          all(float(x.get("salesTaxRate") or 0) >= 1 or float(x.get("value") or 0) == 0
              for x in prev.get("rows", [])),
          f"rates={[x.get('salesTaxRate') for x in prev.get('rows', [])][:5]}")
    check("each row's tax is its own value x its own rate",
          all(near(x["salesTax"], round(x["value"] * x["salesTaxRate"] / 100, 2), 0.02)
              for x in prev.get("rows", [])))

    if not prev.get("canCommit"):
        check("the preview can be committed", False, f"blocking={prev.get('blockingErrors')}")
        return None

    body = {
        "companyId": cid,
        "fileSha256": prev["fileSha256"], "fileName": "stock.xlsx",
        "fileSizeBytes": prev["fileSizeBytes"],
        "asOfDate": date(2026, 7, 1).isoformat(),
        "postInventoryValue": True, "enableInventoryTracking": True,
        "rows": [{"itemName": x["itemName"], "hsCode": x["hsCode"],
                  "isHsCodePartial": x["isHsCodePartial"], "unit": x["unit"],
                  "quantity": x["quantity"], "value": x["value"],
                  "salesTaxRate": x["salesTaxRate"],
                  "lotRefs": x["lotRefs"], "itemTypeId": x["itemTypeId"]}
                 for x in prev["rows"]],
    }
    r = api_call("POST", f"{api}/spreadsheet-import/opening-stock/commit", h, json=body)
    res = r.json() if r.ok else {}
    if not check("the import commits", r.ok, f"http {r.status_code}: {r.text[:200]}"):
        return None
    check("the commit reports the sheet's value",
          near(res.get("totalValueExcludingTax"), e_excl, 0.02),
          f"got {res.get('totalValueExcludingTax')} want {e_excl}")
    check("the commit reports the sheet's sales tax",
          near(res.get("totalSalesTax"), e_tax, 0.5),
          f"got {res.get('totalSalesTax')} want {e_tax}")

    print("\n-- 1b. The stock dashboard reads back the same five figures --")
    t = onhand_totals(api, h, cid)
    if not check("the on-hand feed answers", t is not None):
        return None
    check("dashboard quantity equals the sheet", near(t["qty"], e_qty, QTY_TOL),
          f"got {t['qty']} want {e_qty}")
    check("dashboard excluding tax equals the sheet", near(t["excl"], e_excl, 0.02),
          f"got {t['excl']} want {e_excl}")
    check("dashboard sales tax equals the sheet", near(t["tax"], e_tax, 0.5),
          f"got {t['tax']} want {e_tax}")
    check("dashboard including tax equals the sheet", near(t["incl"], e_incl, 0.6),
          f"got {t['incl']} want {e_incl}")
    check("every item carries a rate", all(float(x["salesTaxRate"]) > 0 for x in t["rows"]),
          f"{sum(1 for x in t['rows'] if not float(x['salesTaxRate']))} without one")
    check("unit cost is value / quantity",
          all(near(x["unitCost"], x["valueExcludingTax"] / x["onHand"], 0.001)
              for x in t["rows"] if float(x["onHand"]) > 0))
    return t


def suite_purchase(api, h, cid, supplier_id, item, before):
    print("\n-- 2. A purchase raises quantity AND value --")
    qty, unit_price, gst = 10.0, 500.0, 18.0
    r = api_call("POST", f"{api}/purchasebills", h, json={
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%dT00:00:00Z"),
        "companyId": cid, "supplierId": supplier_id, "gstRate": gst,
        "items": [{"itemTypeId": item["itemTypeId"], "description": item["itemTypeName"],
                   "quantity": qty, "unit": item.get("uom") or "Pcs",
                   "unitPrice": unit_price}]})
    if not check("the purchase bill is created", r.ok, f"http {r.status_code}: {r.text[:200]}"):
        return None
    after = onhand_row(api, h, cid, item["itemTypeId"])
    check("quantity rises by what was bought",
          near(after["onHand"], float(before["onHand"]) + qty, QTY_TOL),
          f"{before['onHand']} + {qty} != {after['onHand']}")
    check("value rises by quantity x the price paid",
          near(after["valueExcludingTax"],
               float(before["valueExcludingTax"]) + qty * unit_price, 0.02),
          f"expected {float(before['valueExcludingTax']) + qty * unit_price}, "
          f"got {after['valueExcludingTax']}")
    check("the average cost is re-weighted, not replaced",
          near(after["unitCost"],
               (float(before["valueExcludingTax"]) + qty * unit_price)
               / (float(before["onHand"]) + qty), 0.001),
          f"unitCost={after['unitCost']}")
    check("sales tax follows the new value",
          near(after["salesTax"],
               round(float(after["valueExcludingTax"]) * float(after["salesTaxRate"]) / 100, 2), 0.02))
    check("including tax stays excluding + tax",
          near(after["valueIncludingTax"],
               float(after["valueExcludingTax"]) + float(after["salesTax"])))
    return after


def suite_invoice(api, h, cid, client_id, item, before):
    print("\n-- 3. A sale takes value off at COST, not at the sale price --")
    qty = 5.0
    avg = float(before["unitCost"])
    sale_price = avg * 3 + 100            # a fat margin, deliberately
    r = api_call("POST", f"{api}/invoices/standalone", h, json={
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%dT00:00:00Z"),
        "companyId": cid, "clientId": client_id, "gstRate": 18,
        "items": [{"itemTypeId": item["itemTypeId"], "description": item["itemTypeName"],
                   "quantity": qty, "uom": item.get("uom") or "Pcs",
                   "unitPrice": sale_price}]})
    if not check("the invoice is created", r.ok, f"http {r.status_code}: {r.text[:200]}"):
        return None
    after = onhand_row(api, h, cid, item["itemTypeId"])
    check("quantity falls by what was sold",
          near(after["onHand"], float(before["onHand"]) - qty, QTY_TOL),
          f"{before['onHand']} - {qty} != {after['onHand']}")
    check("value falls by quantity x the AVERAGE cost",
          near(after["valueExcludingTax"],
               float(before["valueExcludingTax"]) - qty * avg, 0.02),
          f"expected {float(before['valueExcludingTax']) - qty * avg}, "
          f"got {after['valueExcludingTax']}")
    check("the sale price does NOT drain the stock value",
          not near(after["valueExcludingTax"],
                   float(before["valueExcludingTax"]) - qty * sale_price, 1.0),
          "value moved by the sale price -- cost of goods is being taken at revenue")
    check("the average cost is unchanged by a sale",
          near(after["unitCost"], avg, 0.001),
          f"{avg} became {after['unitCost']}")
    return after


def suite_adjustments(api, h, cid, item, before):
    print("\n-- 4. Adjustments move quantity and value in both directions --")
    today = date.today().isoformat()

    # 4a. Up, with a cost the operator states.
    up_qty, up_cost = 4.0, 750.0
    r = api_call("POST", f"{api}/stock/adjust", h, json={
        "companyId": cid, "itemTypeId": item["itemTypeId"], "delta": up_qty,
        "unitCostExcludingTax": up_cost, "salesTaxRate": before["salesTaxRate"],
        "movementDate": today, "notes": "found in the back room"})
    check("an adjustment up is accepted", r.ok, f"http {r.status_code}: {r.text[:200]}")
    a = onhand_row(api, h, cid, item["itemTypeId"])
    check("adjusting up raises quantity",
          near(a["onHand"], float(before["onHand"]) + up_qty, QTY_TOL), f"{a['onHand']}")
    check("adjusting up at a stated cost raises value by qty x that cost",
          near(a["valueExcludingTax"],
               float(before["valueExcludingTax"]) + up_qty * up_cost, 0.02),
          f"expected {float(before['valueExcludingTax']) + up_qty * up_cost}, "
          f"got {a['valueExcludingTax']}")

    # 4b. Up, with NO cost -- a count correction, valued at the average.
    avg_now = float(a["unitCost"])
    r = api_call("POST", f"{api}/stock/adjust", h, json={
        "companyId": cid, "itemTypeId": item["itemTypeId"], "delta": 2,
        "movementDate": today, "notes": "count correction"})
    check("an adjustment up with no cost is accepted", r.ok, f"http {r.status_code}")
    b = onhand_row(api, h, cid, item["itemTypeId"])
    check("a cost-less adjustment up is valued at the average",
          near(b["valueExcludingTax"], float(a["valueExcludingTax"]) + 2 * avg_now, 0.02),
          f"expected {float(a['valueExcludingTax']) + 2 * avg_now}, got {b['valueExcludingTax']}")
    check("a cost-less adjustment up leaves the average alone",
          near(b["unitCost"], avg_now, 0.001), f"{avg_now} became {b['unitCost']}")

    # 4c. Down -- breakage, always at the average.
    avg_now = float(b["unitCost"])
    r = api_call("POST", f"{api}/stock/adjust", h, json={
        "companyId": cid, "itemTypeId": item["itemTypeId"], "delta": -3,
        "movementDate": today, "notes": "breakage"})
    check("an adjustment down is accepted", r.ok, f"http {r.status_code}: {r.text[:200]}")
    c = onhand_row(api, h, cid, item["itemTypeId"])
    check("adjusting down lowers quantity",
          near(c["onHand"], float(b["onHand"]) - 3, QTY_TOL), f"{c['onHand']}")
    check("adjusting down lowers value at the average",
          near(c["valueExcludingTax"], float(b["valueExcludingTax"]) - 3 * avg_now, 0.02),
          f"expected {float(b['valueExcludingTax']) - 3 * avg_now}, got {c['valueExcludingTax']}")
    check("the average survives an adjustment down",
          near(c["unitCost"], avg_now, 0.001), f"average moved to {c['unitCost']}")

    # 4d. The guard still holds -- no valuing stock that isn't there.
    r = api_call("POST", f"{api}/stock/adjust", h, json={
        "companyId": cid, "itemTypeId": item["itemTypeId"],
        "delta": -(float(c["onHand"]) + 1000), "movementDate": today, "notes": "too much"})
    check("an adjustment below zero is still refused", r.status_code == 400,
          f"http {r.status_code}")
    return c


def suite_feed(api, h, cid, item, grid):
    print("\n-- 5. The movement feed carries the money, row by row --")
    rows = movements(api, h, cid, item["itemTypeId"])
    if not check("the feed returns this item's movements", len(rows) > 0):
        return
    check("every movement is priced", all(float(m.get("unitCost") or 0) > 0 for m in rows),
          f"{sum(1 for m in rows if not float(m.get('unitCost') or 0))} rows without a cost")
    check("each movement's value is its quantity x its unit cost",
          all(near(m["value"], round(float(m["quantity"]) * float(m["unitCost"]), 2), 0.05)
              for m in rows),
          "a row's value does not match its own quantity and cost")
    last = rows[-1]
    check("the last running quantity is the grid's on-hand",
          near(last["runningQuantity"], grid["onHand"], QTY_TOL),
          f"feed {last['runningQuantity']} vs grid {grid['onHand']}")
    check("the last running value is the grid's value",
          near(last["runningValue"], grid["valueExcludingTax"], 0.02),
          f"feed {last['runningValue']} vs grid {grid['valueExcludingTax']}")
    check("ins and outs reconcile with the opening to the grid's quantity",
          near(sum(float(m["quantity"]) * (1 if m["direction"] == "In" else -1) for m in rows)
               + float(grid["openingBalance"]), grid["onHand"], QTY_TOL))
    check("total value in minus total value out is the grid's value",
          near(sum(float(m["value"]) * (1 if m["direction"] == "In" else -1) for m in rows)
               + float(grid["valueIn"]) - float(grid["valueIn"]),
               float(grid["valueExcludingTax"]) - _opening_value(api, h, cid, item), 0.05))


def _opening_value(api, h, cid, item):
    r = api_call("GET", f"{api}/stock/company/{cid}/opening", h)
    if not r.ok:
        return 0.0
    return sum(float(o["valueExcludingTax"]) for o in r.json()
               if o["itemTypeId"] == item["itemTypeId"])


def suite_rate_change(api, h, cid, supplier_id, item):
    print("\n-- 6. A purchase at a different rate re-prices the item --")
    before = onhand_row(api, h, cid, item["itemTypeId"])
    r = api_call("POST", f"{api}/purchasebills", h, json={
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%dT00:00:00Z"),
        "companyId": cid, "supplierId": supplier_id, "gstRate": 25,
        "items": [{"itemTypeId": item["itemTypeId"], "description": item["itemTypeName"],
                   "quantity": 6, "unit": item.get("uom") or "Pcs", "unitPrice": 400}]})
    if not check("a 25% purchase is created", r.ok, f"http {r.status_code}: {r.text[:200]}"):
        return
    after = onhand_row(api, h, cid, item["itemTypeId"])
    check("the item moves onto the new rate", near(after["salesTaxRate"], 25, 0.01),
          f"rate is {after['salesTaxRate']}, was {before['salesTaxRate']}")
    check("sales tax is recomputed at the new rate",
          near(after["salesTax"], round(float(after["valueExcludingTax"]) * 25 / 100, 2), 0.02),
          f"{after['salesTax']}")
    check("including tax follows",
          near(after["valueIncludingTax"],
               float(after["valueExcludingTax"]) + float(after["salesTax"])))


def suite_empty(api, h, cid, client_id, item, grid):
    print("\n-- 7. Selling the last unit leaves value at exactly zero --")
    qty = float(grid["onHand"])
    if qty <= 0:
        skip("stock empties to a value of zero", "nothing left to sell")
        return
    r = api_call("POST", f"{api}/invoices/standalone", h, json={
        "date": datetime.now(timezone.utc).strftime("%Y-%m-%dT00:00:00Z"),
        "companyId": cid, "clientId": client_id, "gstRate": 18,
        "items": [{"itemTypeId": item["itemTypeId"], "description": item["itemTypeName"],
                   "quantity": qty, "uom": item.get("uom") or "Pcs", "unitPrice": 9999}]})
    if not check("the clear-out invoice is created", r.ok,
                 f"http {r.status_code}: {r.text[:200]}"):
        return
    after = onhand_row(api, h, cid, item["itemTypeId"])
    check("quantity is zero", near(after["onHand"], 0, QTY_TOL), f"{after['onHand']}")
    check("value is exactly zero, not a stray paisa",
          float(after["valueExcludingTax"]) == 0.0, f"{after['valueExcludingTax']}")
    check("sales tax on nothing is nothing", float(after["salesTax"]) == 0.0,
          f"{after['salesTax']}")
    check("including tax on nothing is nothing", float(after["valueIncludingTax"]) == 0.0,
          f"{after['valueIncludingTax']}")


# -- Runner -----------------------------------------------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", default="http://localhost:5134")
    ap.add_argument("--username", default="admin")
    ap.add_argument("--password", default="admin123")
    ap.add_argument("--stock-file", default=None,
                    help="a real workbook, imported through the SHIPPED default layout")
    ap.add_argument("--keep", action="store_true")
    args = ap.parse_args()

    api = args.base.rstrip("/") + "/api"
    r = requests.post(f"{api}/auth/login", timeout=60,
                      json={"username": args.username, "password": args.password})
    if not r.ok:
        print(f"FATAL: login failed ({r.status_code} {r.text[:160]})")
        return 2
    h = {"Authorization": f"Bearer {r.json()['token']}"}

    tag = datetime.now().strftime("%m%d%H%M%S")
    r = api_call("POST", f"{api}/companies", h, json={
        "name": f"_stock_valuation {tag}", "brandName": "SVAL",
        "fullAddress": "1 Test Street", "phone": "021-0000000", "ntn": "1234567-8",
        "startingChallanNumber": 1, "startingInvoiceNumber": 1,
        "startingPurchaseBillNumber": 1, "startingGoodsReceiptNumber": 1,
        "startingSalesQuoteNumber": 1, "startingSalesOrderNumber": 1,
        "fbrEnabled": False, "inventoryTrackingEnabled": True, "enableGl": False,
    })
    if not r.ok:
        print(f"FATAL: create company failed ({r.status_code} {r.text[:200]})")
        return 2
    cid = r.json()["id"]
    print(f"\nCompany id={cid}  base={args.base}")

    api_call("POST", f"{api}/accounts/company/{cid}/seed-wholesale", h)
    api_call("POST", f"{api}/stock/company/{cid}/flow-version", h, json={"version": 2})

    r = api_call("POST", f"{api}/clients", h, json={
        "name": f"Valuation Client {tag}", "address": "1 Test Road, Karachi",
        "phone": "021-1234567", "companyId": cid, "registrationType": "Unregistered"})
    client_id = r.json()["id"] if r.ok else None
    r = api_call("POST", f"{api}/suppliers", h, json={
        "name": f"Valuation Supplier {tag}", "companyId": cid,
        "registrationType": "Unregistered"})
    supplier_id = r.json()["id"] if r.ok else None
    if not client_id or not supplier_id:
        print("FATAL: could not create the client / supplier")
        return 2

    try:
        if args.stock_file:
            path = os.path.abspath(args.stock_file)
            if not os.path.exists(path):
                print(f"FATAL: --stock-file not found: {path}")
                return 2
            with open(path, "rb") as f:
                book = f.read()
            # No mapping: the SHIPPED default layout has to recognise the file.
            grid = suite_import(api, h, cid, book, None, read_real_sheet(path),
                                f"real sheet, shipped layout -- {os.path.basename(path)}")
        else:
            rows = sheet_rows(tag)
            grid = suite_import(api, h, cid, stock_workbook(rows), STOCK_MAPPING,
                                sheet_totals(rows), "synthetic sheet")
        if grid is None:
            print("\nImport failed -- the later suites depend on it, stopping here.")
            return report()

        # Work the most valuable item, so a sale never runs the bin dry early.
        item = max(grid["rows"], key=lambda x: float(x["valueExcludingTax"]))
        print(f"\nWorking item: {item['itemTypeName']} "
              f"(qty {item['onHand']}, value {item['valueExcludingTax']}, "
              f"rate {item['salesTaxRate']}%)")

        after_buy = suite_purchase(api, h, cid, supplier_id, item, item)
        if after_buy is None:
            return report()
        after_sell = suite_invoice(api, h, cid, client_id, item, after_buy)
        if after_sell is None:
            return report()
        after_adj = suite_adjustments(api, h, cid, item, after_sell)
        suite_feed(api, h, cid, item, after_adj)
        suite_rate_change(api, h, cid, supplier_id, item)
        suite_empty(api, h, cid, client_id, item,
                    onhand_row(api, h, cid, item["itemTypeId"]))
    finally:
        if args.keep:
            print(f"\nKeeping company id={cid} (--keep)")
        else:
            d = api_call("DELETE", f"{api}/companies/{cid}", h)
            print(f"\nTeardown: delete company returned {d.status_code}")

    return report()


def report():
    print("\n" + "=" * 72)
    p = sum(1 for s, _, _ in results if s == PASS)
    f = sum(1 for s, _, _ in results if s == FAIL)
    s = sum(1 for st, _, _ in results if st == SKIP)
    for st, name, detail in results:
        if st == FAIL:
            print(f"  FAIL  {name} -- {detail}")
    print(f"{p}/{p + f} checks passed" + (f", {s} skipped" if s else ""))
    print("=" * 72)
    return 0 if f == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
